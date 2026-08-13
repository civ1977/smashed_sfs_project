from datetime import date, timedelta
from io import BytesIO

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse

from accounts.models import Teacher
from grades.models import Grade, SubjectMapping
from students.models import Section, Student
from .views import (
    _age_from_birthday,
    _build_subject_rows,
    _final_average,
    _gate_finals_pending_term3,
    _remarks_for,
    _split_core_elective,
    _student_display_name,
)


class FinalAverageTests(TestCase):
    def test_all_three_terms_present(self):
        self.assertEqual(_final_average([80, 85, 90]), 85)

    def test_rounds_to_nearest_int(self):
        self.assertEqual(_final_average([80, 81, 81]), round((80 + 81 + 81) / 3))

    def test_missing_terms_are_ignored(self):
        self.assertEqual(_final_average([80, None, None]), 80)

    def test_no_terms_returns_none(self):
        self.assertIsNone(_final_average([None, None, None]))


class RemarksForTests(TestCase):
    def test_none_final_has_no_remark(self):
        self.assertIsNone(_remarks_for(None))

    def test_75_is_passed(self):
        self.assertEqual(_remarks_for(75), 'Passed')

    def test_74_is_failed(self):
        self.assertEqual(_remarks_for(74), 'Failed')


class GateFinalsPendingTerm3Tests(TestCase):
    def test_all_subjects_complete_returns_true(self):
        rows = [
            {'term_3': 90, 'final': 85, 'remarks': 'Passed'},
            {'term_3': 88, 'final': 80, 'remarks': 'Passed'},
        ]
        self.assertTrue(_gate_finals_pending_term3(rows))
        # Untouched when already complete.
        self.assertEqual(rows[0]['final'], 85)

    def test_one_subject_missing_term3_blanks_its_final_and_returns_false(self):
        rows = [
            {'term_3': 90, 'final': 85, 'remarks': 'Passed'},
            {'term_3': None, 'final': 80, 'remarks': 'Passed'},
        ]
        complete = _gate_finals_pending_term3(rows)
        self.assertFalse(complete)
        self.assertIsNone(rows[1]['final'])
        self.assertIsNone(rows[1]['remarks'])
        # A subject that does have Term 3 data is left alone even though
        # the report as a whole isn't complete yet.
        self.assertEqual(rows[0]['final'], 85)


class SplitCoreElectiveTests(TestCase):
    def test_splits_by_is_elective_flag(self):
        rows = [
            {'subject_name': 'Math', 'is_elective': False},
            {'subject_name': 'Research', 'is_elective': True},
        ]
        core, elective = _split_core_elective(rows)
        self.assertEqual([r['subject_name'] for r in core], ['Math'])
        self.assertEqual([r['subject_name'] for r in elective], ['Research'])


class AgeFromBirthdayTests(TestCase):
    def test_none_birthday_returns_none(self):
        self.assertIsNone(_age_from_birthday(None))

    def test_birthday_already_passed_this_year(self):
        today = date.today()
        birthday = today.replace(year=today.year - 20) - timedelta(days=1)
        self.assertEqual(_age_from_birthday(birthday), 20)

    def test_birthday_not_yet_reached_this_year(self):
        today = date.today()
        birthday = today.replace(year=today.year - 20) + timedelta(days=1)
        self.assertEqual(_age_from_birthday(birthday), 19)


class BuildSubjectRowsTests(TestCase):
    def setUp(self):
        self.student = Student.objects.create(
            lrn='123456789012', surname='Dela Cruz', name='Juan', sex='MALE',
            school_g10='Test JHS', school_address_g10='Test Address',
            average_g10='90.00', section_id=1, adviser_id=1,
        )
        self.math = SubjectMapping.objects.create(
            school_profile_id=1, grade_level='11', strand='', subject_number=1,
            subject_name='Math', created_by=1, is_elective=False,
        )
        self.research = SubjectMapping.objects.create(
            school_profile_id=1, grade_level='11', strand='', subject_number=2,
            subject_name='Research', created_by=1, is_elective=True,
        )

    def test_groups_by_subject_and_computes_final(self):
        Grade.objects.create(lrn=self.student.lrn, mapping_id=self.math.mapping_id, term=1, grade=80, uploaded_by=1)
        Grade.objects.create(lrn=self.student.lrn, mapping_id=self.math.mapping_id, term=2, grade=84, uploaded_by=1)
        Grade.objects.create(lrn=self.student.lrn, mapping_id=self.math.mapping_id, term=3, grade=88, uploaded_by=1)

        rows = _build_subject_rows(self.student.lrn)

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row['subject_name'], 'Math')
        self.assertEqual(row['final'], 84)
        self.assertEqual(row['remarks'], 'Passed')

    def test_rows_are_ordered_by_subject_number(self):
        Grade.objects.create(lrn=self.student.lrn, mapping_id=self.research.mapping_id, term=1, grade=80, uploaded_by=1)
        Grade.objects.create(lrn=self.student.lrn, mapping_id=self.math.mapping_id, term=1, grade=90, uploaded_by=1)

        rows = _build_subject_rows(self.student.lrn)

        self.assertEqual([r['subject_name'] for r in rows], ['Math', 'Research'])

    def test_grade_level_filter_excludes_other_levels(self):
        g12_subject = SubjectMapping.objects.create(
            school_profile_id=1, grade_level='12', strand='', subject_number=1,
            subject_name='Practical Research 2', created_by=1,
        )
        Grade.objects.create(lrn=self.student.lrn, mapping_id=self.math.mapping_id, term=1, grade=80, uploaded_by=1)
        Grade.objects.create(lrn=self.student.lrn, mapping_id=g12_subject.mapping_id, term=1, grade=85, uploaded_by=1)

        rows = _build_subject_rows(self.student.lrn, grade_level='11')

        self.assertEqual([r['subject_name'] for r in rows], ['Math'])


class StudentDisplayNameTests(TestCase):
    def test_includes_middle_initial_with_period(self):
        student = Student(surname='Dela Cruz', name='Juan', middle_name='Santos')
        self.assertEqual(_student_display_name(student), 'Dela Cruz, Juan S.')

    def test_omits_middle_initial_when_blank(self):
        student = Student(surname='Dela Cruz', name='Juan', middle_name='')
        self.assertEqual(_student_display_name(student), 'Dela Cruz, Juan')

    def test_omits_middle_initial_when_none(self):
        student = Student(surname='Dela Cruz', name='Juan', middle_name=None)
        self.assertEqual(_student_display_name(student), 'Dela Cruz, Juan')


class SummaryOfRatingsViewTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username='adviser1', password='testpass123')
        self.teacher = Teacher.objects.create(
            user=user, username='adviser1', password='testpass123',
            full_name='Adviser One', position='Teacher I', role=Teacher.ROLE_ADVISER,
        )
        self.section = Section.objects.create(
            grade_level='11', track='Academic', strand='STEM', section_name='Newton',
            modality='Face-to-face', adviser_id=self.teacher.teacher_id, school_profile_id=1,
        )
        self.math = SubjectMapping.objects.create(
            school_profile_id=1, grade_level='11', strand='', subject_number=1,
            subject_name='Math', created_by=self.teacher.teacher_id,
        )
        self.research = SubjectMapping.objects.create(
            school_profile_id=1, grade_level='11', strand='', subject_number=2,
            subject_name='Research', created_by=self.teacher.teacher_id,
        )
        self.student_a = Student.objects.create(
            lrn='111111111111', surname='Aquino', name='Ana', sex='FEMALE',
            school_g10='Test JHS', school_address_g10='Test Address', average_g10='90.00',
            section_id=self.section.section_id, adviser_id=self.teacher.teacher_id,
        )
        self.student_b = Student.objects.create(
            lrn='222222222222', surname='Bautista', name='Ben', sex='MALE',
            school_g10='Test JHS', school_address_g10='Test Address', average_g10='88.00',
            section_id=self.section.section_id, adviser_id=self.teacher.teacher_id,
        )
        self.client.force_login(user)

    def test_columns_align_across_students_with_different_subjects(self):
        # Student A has grades in both subjects; Student B only has Math -
        # Research's column should still render as blank for B rather than
        # shifting B's Math value into Research's column.
        for term in (1, 2, 3):
            Grade.objects.create(lrn=self.student_a.lrn, mapping_id=self.math.mapping_id, term=term, grade=80, uploaded_by=self.teacher.teacher_id)
            Grade.objects.create(lrn=self.student_a.lrn, mapping_id=self.research.mapping_id, term=term, grade=90, uploaded_by=self.teacher.teacher_id)
            Grade.objects.create(lrn=self.student_b.lrn, mapping_id=self.math.mapping_id, term=term, grade=85, uploaded_by=self.teacher.teacher_id)

        response = self.client.get(reverse('summary_of_ratings'))

        self.assertEqual(response.status_code, 200)
        student_rows = {row['student'].lrn: row for row in response.context['student_rows']}

        row_a = student_rows[self.student_a.lrn]
        self.assertEqual(row_a['subject_cells'][0]['final'], 80)
        self.assertEqual(row_a['subject_cells'][1]['final'], 90)
        self.assertEqual(row_a['general_average'], 85)

        row_b = student_rows[self.student_b.lrn]
        self.assertEqual(row_b['subject_cells'][0]['final'], 85)
        self.assertIsNone(row_b['subject_cells'][1]['final'])
        # Research has no grades at all for B, so the general average isn't
        # gated by it - only subjects B actually has grades in count.
        self.assertEqual(row_b['general_average'], 85)

    def test_no_advisory_section_redirects_with_message(self):
        Section.objects.filter(section_id=self.section.section_id).delete()

        response = self.client.get(reverse('summary_of_ratings'))

        self.assertRedirects(response, reverse('select_student_report'))

    def test_students_sorted_all_male_then_all_female(self):
        # Student A (Aquino, FEMALE) sorts before Student B (Bautista, MALE)
        # alphabetically - the view must override that with sex first.
        response = self.client.get(reverse('summary_of_ratings'))

        lrns_in_order = [row['student'].lrn for row in response.context['student_rows']]

        self.assertEqual(lrns_in_order, [self.student_b.lrn, self.student_a.lrn])

    def test_page_renders_frozen_header_and_column_markup(self):
        response = self.client.get(reverse('summary_of_ratings'))

        self.assertContains(response, 'col-frozen')
        self.assertContains(response, 'sor-row1-height')

    def test_export_produces_xlsx_matching_the_page_data(self):
        for term in (1, 2, 3):
            Grade.objects.create(lrn=self.student_a.lrn, mapping_id=self.math.mapping_id, term=term, grade=80, uploaded_by=self.teacher.teacher_id)
            Grade.objects.create(lrn=self.student_b.lrn, mapping_id=self.math.mapping_id, term=term, grade=85, uploaded_by=self.teacher.teacher_id)

        response = self.client.get(reverse('export_summary_of_ratings'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        self.assertEqual(ws['A1'].value, 'Student Name')
        self.assertEqual(ws['B1'].value, 'Math')
        self.assertEqual(ws['B2'].value, 'Term 1')
        # Research has no grades for either student in this test, so its
        # header still occupies its own 4 columns (F-I) ahead of General
        # Average in J - column alignment must survive into the export too.
        self.assertEqual(ws['F1'].value, 'Research')
        self.assertEqual(ws['J1'].value, 'General Average')

        # Row 3 is Student B (MALE, sorted first); row 4 is Student A (FEMALE).
        self.assertEqual(ws['A3'].value, 'Bautista, Ben')
        self.assertEqual(ws['E3'].value, 85)
        self.assertEqual(ws['A4'].value, 'Aquino, Ana')
        self.assertEqual(ws['E4'].value, 80)

        self.assertEqual(ws.freeze_panes, 'B3')

    def test_page_shows_middle_initial_with_period(self):
        self.student_a.middle_name = 'Reyes'
        self.student_a.save()

        response = self.client.get(reverse('summary_of_ratings'))

        self.assertContains(response, 'Aquino, Ana R.')
