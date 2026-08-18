from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.urls import reverse
from django.utils.html import format_html

from accounts.models import Teacher
from students.models import Student, SchoolProfile, Section
from grades.models import Grade, SubjectMapping, StudentTermRemark, ATTENDANCE_MONTHS, AttendanceMark, SchoolCalendarException, TeacherSubjectAssignment, SubjectTestMaxScore
from grades.views import MONTH_NAMES, _school_days_in_month, build_student_grade_sheet, _score_stats, WEEKDAYS, round_half_up


def _get_teacher(request):
    return Teacher.objects.get(username=request.user.username)


def _final_average(term_grades):
    valid = [g for g in term_grades if g is not None]
    return round_half_up(sum(valid) / len(valid)) if valid else None


def _remarks_for(final):
    if final is None:
        return None
    return 'Passed' if final >= 75 else 'Failed'


def _student_display_name(student):
    """'Surname, Name M.' - middle name reduced to its initial plus a
    period; omitted entirely for students with no middle name on file
    rather than showing a dangling comma or period."""
    name = f'{student.surname.strip()}, {student.name.strip()}'
    if student.middle_name and student.middle_name.strip():
        name += f' {student.middle_name.strip()[0].upper()}.'
    return name


def _build_subject_rows(lrn, grade_level=None, term=None):
    """Build per-subject grade rows for a student, optionally scoped to one
    subject grade_level (a student's Grade table can hold both Grade 11 and
    Grade 12 rows once they've been promoted) and/or one term (for a
    single-quarter SF9 export)."""
    grades = Grade.objects.filter(lrn=lrn)
    if term is not None:
        grades = grades.filter(term=term)

    mapping_ids = {g.mapping_id for g in grades}
    mappings = {m.mapping_id: m for m in SubjectMapping.objects.filter(mapping_id__in=mapping_ids)}

    grouped = {}
    for grade in grades:
        mapping = mappings.get(grade.mapping_id)
        if grade_level is not None and (mapping is None or mapping.grade_level != grade_level):
            continue
        grouped.setdefault(grade.mapping_id, {1: None, 2: None, 3: None})[grade.term] = grade.grade

    subject_rows = []
    for mapping_id, terms in grouped.items():
        mapping = mappings.get(mapping_id)
        subject_name = mapping.subject_name if mapping else f'Subject {mapping_id}'
        term_grades = [terms[1], terms[2], terms[3]]
        final = _final_average(term_grades)
        subject_rows.append({
            'subject_name': subject_name,
            'mapping_id': mapping_id,
            # Same field the class gradesheet orders by and the adviser's
            # own "Arrange Subjects" tool writes to (grades/views.py's
            # arrange_subjects) - one shared order across gradesheet,
            # SF9, and SF10 instead of each picking its own sort.
            'subject_number': mapping.subject_number if mapping else 0,
            'is_elective': mapping.is_elective if mapping else False,
            'term_1': terms[1],
            'term_2': terms[2],
            'term_3': terms[3],
            'final': final,
            'remarks': _remarks_for(final),
        })
    subject_rows.sort(key=lambda row: (row['subject_number'], row['subject_name']))
    return subject_rows


def _grade_levels_for_student(lrn):
    mapping_ids = Grade.objects.filter(lrn=lrn).values_list('mapping_id', flat=True)
    levels = SubjectMapping.objects.filter(mapping_id__in=set(mapping_ids)).values_list('grade_level', flat=True)
    return sorted(set(levels))


MONTH_NUMBERS = {
    'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12,
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4,
}


def _school_year_start():
    """First calendar year of the school year containing today - e.g. 2026
    for any date from Jun 2026 through Apr 2027 - so Jun-Dec months below
    resolve to that year and Jan-Apr months resolve to the year after.
    Matches the same today-anchored convention attendance_grid/SF2 already
    use (grades/views.py's attendance_grid defaults its year/month to
    date.today()); SchoolProfile.school_year is a free-text display label
    only (see its one other use at reports/views.py's SF2 export, cell
    U8) and isn't a reliable source for real calendar-date math - trusting
    it here previously caused SF9 to silently look up the wrong calendar
    year whenever that label drifted from the dates attendance was
    actually being recorded against, so it's deliberately not read here."""
    today = date.today()
    return today.year if today.month >= 6 else today.year - 1


TERM_MONTHS = {
    1: ['Jun', 'Jul', 'Aug'],
    2: ['Sep', 'Oct', 'Nov', 'Dec'],
    3: ['Jan', 'Feb', 'Mar', 'Apr'],
}


def _visible_attendance_months(lrn):
    """SF9's Attendance Record only shows the months for terms that already
    have grade data - Term 1's Jun-Aug always show; Sep-Dec (Term 2) and
    Jan-Apr (Term 3) only appear once that term has at least one grade
    entered for this student, the same signal _gate_finals_pending_term3
    uses to gate Final Grade/Remarks."""
    months = list(TERM_MONTHS[1])
    if Grade.objects.filter(lrn=lrn, term=2).exists():
        months += TERM_MONTHS[2]
    if Grade.objects.filter(lrn=lrn, term=3).exists():
        months += TERM_MONTHS[3]
    return months


def _attendance_rows(lrn, school_profile_id):
    """Per-month class days/present/absent for the SF9/SF10 attendance
    grids, derived from real AttendanceMark exceptions and the school
    calendar - a school day with no mark is Present, matching the SF2
    form's own convention (see _fill_sf2_gender_block below). The Month/
    Total header row always covers all 11 months - only the data cells
    for months the current term progress doesn't allow showing yet (see
    _visible_attendance_months) are left blank; April is further capped
    to its first 8 days once Term 3 unlocks it, rather than the full
    month."""
    start_year = _school_year_start()
    visible_months = _visible_attendance_months(lrn)
    rows = []
    for month in ATTENDANCE_MONTHS:
        if month not in visible_months:
            rows.append({'month': month, 'total': None, 'present': None, 'absent': None})
            continue
        month_num = MONTH_NUMBERS[month]
        year = start_year if month_num >= 6 else start_year + 1
        school_days = _school_days_in_month(school_profile_id, year, month_num)
        if month == 'Apr':
            school_days = [d for d in school_days if d.day <= 8]
        class_days = len(school_days)
        if class_days:
            absent = AttendanceMark.objects.filter(
                lrn=lrn, date__in=school_days, status=AttendanceMark.STATUS_ABSENT,
            ).count()
            present = class_days - absent
        else:
            absent = None
            present = None
        rows.append({
            'month': month,
            'total': class_days or None,
            'present': present,
            'absent': absent,
        })
    return rows


def _attendance_totals(attendance_rows):
    """Year-to-date totals for the SF9 monthly attendance grid."""
    known_rows = [r for r in attendance_rows if r['total'] is not None]
    if not known_rows:
        return {'class_days': None, 'present': None, 'absent': None}
    return {
        'class_days': sum(r['total'] for r in known_rows),
        'present': sum(r['present'] for r in known_rows),
        'absent': sum(r['absent'] for r in known_rows),
    }


def _age_from_birthday(birthday):
    if not birthday:
        return None
    today = date.today()
    had_birthday_this_year = (today.month, today.day) >= (birthday.month, birthday.day)
    return today.year - birthday.year - (0 if had_birthday_this_year else 1)


def _gate_finals_pending_term3(subject_rows):
    """A subject's Final Grade/Remarks only show once Term 3 data has
    actually been uploaded for it. Returns True if every subject on this
    report has a displayable final, which gates whether the General
    Average shows."""
    all_complete = True
    for row in subject_rows:
        if row['term_3'] is None:
            row['final'] = None
            row['remarks'] = None
            all_complete = False
    return all_complete


def _split_core_elective(subject_rows):
    """SF9 groups Learning Areas under 'Core Subjects' and 'Elective Subjects'
    headers, using SubjectMapping.is_elective - set from the adviser's own
    "Arrange Subjects" Core/Elective drag-and-drop page (grades/views.py's
    arrange_subjects), not guessed from the subject's name."""
    core_rows = [r for r in subject_rows if not r['is_elective']]
    elective_rows = [r for r in subject_rows if r['is_elective']]
    return core_rows, elective_rows


@login_required
def select_student_for_report(request):
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    return render(request, 'reports/select_student.html', {
        'is_subject_teacher': teacher.role == Teacher.ROLE_SUBJECT_TEACHER,
    })


@login_required
def subject_statistics_report(request):
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
    if not section:
        messages.error(request, 'No advisory section found for your account.')
        return redirect('select_student_report')

    student_lrns = list(Student.objects.filter(adviser_id=teacher.teacher_id).values_list('lrn', flat=True))

    subjects = SubjectMapping.objects.filter(
        school_profile_id=section.school_profile_id,
        grade_level=section.grade_level,
    ).order_by('subject_name')

    # Highest-possible-score settings live per teaching assignment, not per
    # subject - find whichever assignment covers this section for each
    # subject so Pre-Test/Final Exam MPS can use the right denominator.
    assignment_by_mapping = {
        a.mapping_id: a
        for a in TeacherSubjectAssignment.objects.filter(section_id=section.section_id)
    }
    max_scores = {
        (m.assignment_id, m.term): m
        for m in SubjectTestMaxScore.objects.filter(
            assignment_id__in=[a.assignment_id for a in assignment_by_mapping.values()]
        )
    }

    subject_scores = []
    for subject in subjects:
        grades = Grade.objects.filter(lrn__in=student_lrns, mapping_id=subject.mapping_id)
        final_by_term = {1: [], 2: [], 3: []}
        pretest_by_term = {1: [], 2: [], 3: []}
        final_exam_by_term = {1: [], 2: [], 3: []}
        for g in grades:
            if g.grade is not None:
                final_by_term[g.term].append(g.grade)
            if g.pretest_score is not None:
                pretest_by_term[g.term].append(g.pretest_score)
            if g.final_exam_score is not None:
                final_exam_by_term[g.term].append(g.final_exam_score)

        subject_scores.append({
            'subject': subject,
            'assignment': assignment_by_mapping.get(subject.mapping_id),
            'final_by_term': final_by_term,
            'pretest_by_term': pretest_by_term,
            'final_exam_by_term': final_exam_by_term,
        })

    # Built as three independent term tables (rather than one dict keyed by
    # term) so the template can loop with plain dotted access instead of
    # int-keyed dict lookups.
    terms = []
    for term in (1, 2, 3):
        rows = []
        for entry in subject_scores:
            assignment = entry['assignment']
            max_score = max_scores.get((assignment.assignment_id, term)) if assignment else None
            rows.append({
                'subject': entry['subject'],
                'final_rating': _score_stats(entry['final_by_term'][term], 100),
                'pretest': _score_stats(entry['pretest_by_term'][term], max_score.pretest_max if max_score else None),
                'final_exam': _score_stats(entry['final_exam_by_term'][term], max_score.final_exam_max if max_score else None),
            })
        terms.append({'term': term, 'rows': rows})

    return render(request, 'reports/subject_statistics.html', {
        'section': section,
        'terms': terms,
    })


@login_required
def view_student_ratings(request, student_lrn):
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    student = get_object_or_404(Student, lrn=student_lrn, adviser_id=teacher.teacher_id)
    subject_grades, subject_names, general_average, grades = build_student_grade_sheet(student_lrn)

    return render(request, 'reports/student_ratings.html', {
        'student': student,
        'subject_grades': subject_grades,
        'subject_names': subject_names,
        'general_average': general_average,
        'grades': grades,
    })


@login_required
def report_cards(request):
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    students = Student.objects.filter(adviser_id=teacher.teacher_id).order_by('surname', 'name')
    males = [s for s in students if s.sex in ('MALE', 'M')]
    females = [s for s in students if s.sex in ('FEMALE', 'F')]
    students = males + females

    return render(request, 'reports/report_cards.html', {'students': students})


def _summary_of_ratings_data(teacher):
    """Shared data-shaping for the Summary of Ratings page and its Excel
    export - one advisory section's students (all-MALE-then-all-FEMALE),
    aligned to a shared subject column order. Returns (section, subjects,
    student_rows); section is None (subjects/student_rows empty) if the
    teacher has no advisory section."""
    section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
    if not section:
        return None, [], []

    students = Student.objects.filter(adviser_id=teacher.teacher_id).order_by('surname', 'name')
    males = [s for s in students if s.sex in ('MALE', 'M')]
    females = [s for s in students if s.sex in ('FEMALE', 'F')]
    students = males + females

    # Master subject list/order for the section's grade level - every
    # student's row is aligned to these same columns (via mapping_id) below,
    # same source SF9/subject_statistics use, so a subject a given student
    # has no grades for yet still gets a blank cell in the right column
    # instead of shifting the row.
    subjects = list(SubjectMapping.objects.filter(
        school_profile_id=section.school_profile_id,
        grade_level=section.grade_level,
    ).order_by('subject_number', 'subject_name'))

    student_rows = []
    for student in students:
        subject_rows = _build_subject_rows(student.lrn, grade_level=section.grade_level)
        all_finals_ready = _gate_finals_pending_term3(subject_rows)
        rows_by_mapping = {row['mapping_id']: row for row in subject_rows}

        empty_cell = {'term_1': None, 'term_2': None, 'term_3': None, 'final': None}
        subject_cells = [rows_by_mapping.get(subject.mapping_id, empty_cell) for subject in subjects]

        finals = [row['final'] for row in subject_rows if row['final'] is not None]
        general_average = round_half_up(sum(finals) / len(finals)) if (all_finals_ready and finals) else None

        student_rows.append({
            'student': student,
            'display_name': _student_display_name(student),
            'subject_cells': subject_cells,
            'general_average': general_average,
        })

    return section, subjects, student_rows


@login_required
def summary_of_ratings(request):
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    section, subjects, student_rows = _summary_of_ratings_data(teacher)
    if not section:
        messages.error(request, 'No advisory section found for your account.')
        return redirect('select_student_report')

    return render(request, 'reports/summary_of_ratings.html', {
        'section': section,
        'subjects': subjects,
        'student_rows': student_rows,
    })


@login_required
def export_summary_of_ratings(request):
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    section, subjects, student_rows = _summary_of_ratings_data(teacher)
    if not section:
        messages.error(request, 'No advisory section found for your account.')
        return redirect('select_student_report')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Summary of Ratings'

    header_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')

    ws.cell(row=1, column=1, value='Student Name')
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    for subject in subjects:
        ws.cell(row=1, column=col, value=subject.subject_name)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        for offset, label in enumerate(('Term 1', 'Term 2', 'Term 3', 'Average')):
            ws.cell(row=2, column=col + offset, value=label)
        col += 4

    last_col = col
    ws.cell(row=1, column=last_col, value='General Average')
    ws.merge_cells(start_row=1, start_column=last_col, end_row=2, end_column=last_col)

    for header_row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=last_col):
        for cell in header_row:
            cell.font = header_font
            cell.alignment = center

    row_num = 3
    for row_data in student_rows:
        ws.cell(row=row_num, column=1, value=row_data['display_name'])
        col = 2
        for cell_data in row_data['subject_cells']:
            ws.cell(row=row_num, column=col, value=cell_data['term_1'])
            ws.cell(row=row_num, column=col + 1, value=cell_data['term_2'])
            ws.cell(row=row_num, column=col + 2, value=cell_data['term_3'])
            ws.cell(row=row_num, column=col + 3, value=cell_data['final'])
            col += 4
        ws.cell(row=row_num, column=last_col, value=row_data['general_average'])
        row_num += 1

    # Excel's real Freeze Panes feature (View > Freeze Panes), matching the
    # page's CSS sticky header/first-column: everything above row 3 and
    # left of column B stays fixed while scrolling.
    ws.freeze_panes = 'B3'

    ws.column_dimensions['A'].width = 28
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Summary_of_Ratings_{section.section_name}.xlsx'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def view_sf9(request, student_lrn):
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    student = get_object_or_404(Student, lrn=student_lrn, adviser_id=teacher.teacher_id)

    school_profile = None
    if teacher.school_profile_id:
        school_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()

    section = Section.objects.filter(section_id=student.section_id).first()

    subject_rows = _build_subject_rows(student_lrn, grade_level=section.grade_level if section else None)
    all_finals_ready = _gate_finals_pending_term3(subject_rows)
    core_rows, elective_rows = _split_core_elective(subject_rows)

    finals = [row['final'] for row in subject_rows if row['final'] is not None]
    general_average = round_half_up(sum(finals) / len(finals)) if (all_finals_ready and finals) else None

    school_profile_id = section.school_profile_id if section else teacher.school_profile_id
    attendance_rows = _attendance_rows(student_lrn, school_profile_id)
    attendance_by_month = {row['month']: row for row in attendance_rows}

    term_remarks = {r.term: r.remark for r in StudentTermRemark.objects.filter(lrn=student_lrn)}

    return render(request, 'reports/sf9.html', {
        'student': student,
        'school_profile': school_profile,
        'section': section,
        'age': _age_from_birthday(student.birthday),
        'core_rows': core_rows,
        'elective_rows': elective_rows,
        'general_average': general_average,
        'term_remarks': term_remarks,
        'attendance_rows': attendance_rows,
        'attendance_by_month': attendance_by_month,
        'attendance_totals': _attendance_totals(attendance_rows),
        'attendance_months': ATTENDANCE_MONTHS,
        'attendance_complete': all(row['total'] is not None for row in attendance_rows),
        'teacher': teacher,
    })


@login_required
def view_sf10(request, student_lrn):
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    student = get_object_or_404(Student, lrn=student_lrn, adviser_id=teacher.teacher_id)

    school_profile = None
    if teacher.school_profile_id:
        school_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()

    section = Section.objects.filter(section_id=student.section_id).first()

    grade_level_sections = []
    for grade_level in _grade_levels_for_student(student_lrn):
        subject_rows = _build_subject_rows(student_lrn, grade_level=grade_level)
        all_finals_ready = _gate_finals_pending_term3(subject_rows)
        finals = [row['final'] for row in subject_rows if row['final'] is not None]
        general_average = round_half_up(sum(finals) / len(finals), 2) if (all_finals_ready and finals) else None
        core_rows, elective_rows = _split_core_elective(subject_rows)
        grade_level_sections.append({
            'grade_level': grade_level,
            'subject_rows': subject_rows,
            'core_rows': core_rows,
            'elective_rows': elective_rows,
            'general_average': general_average,
            'remarks': _remarks_for(general_average) if all_finals_ready else None,
        })

    return render(request, 'reports/sf10.html', {
        'student': student,
        'school_profile': school_profile,
        'section': section,
        'grade_level_sections': grade_level_sections,
        'attendance_rows': _attendance_rows(
            student_lrn,
            section.school_profile_id if section else teacher.school_profile_id,
        ),
    })


def _format_schedule(schedule):
    """Render a SubjectTestMaxScore.schedule JSON dict
    ({"Mon": {"start": "08:00", "end": "09:00"}, ...}) as "Mon 08:00-09:00,
    Wed 08:00-09:00" for display. Returns '—' when nothing is set."""
    if not schedule:
        return '—'
    parts = []
    for day in WEEKDAYS:
        entry = schedule.get(day)
        if not entry:
            continue
        start = entry.get('start') or '?'
        end = entry.get('end') or '?'
        parts.append(f'{day} {start}-{end}')
    return ', '.join(parts) if parts else '—'


def _format_schedule_all_terms(assignment_id):
    """A per-assignment schedule summary across all 3 terms (schedule is
    now stored per SubjectTestMaxScore row, i.e. per assignment+term, since
    it can differ term to term) - e.g. "T1: Mon 08:00-09:00 | T2: — | T3:
    Mon 09:00-10:00"."""
    max_scores_by_term = {
        m.term: m for m in SubjectTestMaxScore.objects.filter(assignment_id=assignment_id)
    }
    return ' | '.join(
        f'T{term}: {_format_schedule(max_scores_by_term.get(term).schedule if max_scores_by_term.get(term) else None)}'
        for term in (1, 2, 3)
    )


@login_required
def view_sf7(request):
    """SF7 - School Personnel Assignment List and Basic Profile: only the
    requesting user's own record - their position, advisory section (if
    any), and subject-teaching assignments. Not tied to a single student,
    unlike SF9/SF10."""
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    school_profile = None
    if teacher.school_profile_id:
        school_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()

    personnel = [teacher]
    personnel_ids = [p.teacher_id for p in personnel]

    advisory_sections = {
        s.adviser_id: s for s in Section.objects.filter(adviser_id__in=personnel_ids)
    }

    assignments_by_teacher = {}
    for a in TeacherSubjectAssignment.objects.filter(teacher_id__in=personnel_ids):
        assignments_by_teacher.setdefault(a.teacher_id, []).append(a)

    section_ids = {a.section_id for assignments in assignments_by_teacher.values() for a in assignments}
    mapping_ids = {a.mapping_id for assignments in assignments_by_teacher.values() for a in assignments}
    sections_by_id = {s.section_id: s for s in Section.objects.filter(section_id__in=section_ids)}
    mappings_by_id = {m.mapping_id: m for m in SubjectMapping.objects.filter(mapping_id__in=mapping_ids)}

    rows = []
    for person in personnel:
        assignment_rows = []
        for a in assignments_by_teacher.get(person.teacher_id, []):
            section = sections_by_id.get(a.section_id)
            mapping = mappings_by_id.get(a.mapping_id)
            assignment_rows.append({
                'subject_name': mapping.subject_name if mapping else 'Unknown subject',
                'section_label': f'{section.grade_level}-{section.strand}-{section.section_name}' if section else 'Unknown section',
                'schedule': _format_schedule_all_terms(a.assignment_id),
            })

        rows.append({
            'teacher': person,
            'advisory_section': advisory_sections.get(person.teacher_id),
            'assignments': assignment_rows,
        })

    return render(request, 'reports/sf7.html', {
        'school_profile': school_profile,
        'rows': rows,
    })


# =====================================================================
# SF2 (Daily Attendance Report) Excel export
#
# SF2.xlsx (reports/xlsx_templates/SF2.xlsx) is used purely as a layout
# template - it was exported out of DepEd's LIS and most of its formulas
# are broken external-workbook links (`[1]INPUT_Homeroom_Data!...` etc.)
# that reference a file we don't have. Every cell we touch below is
# overwritten with a real, static value computed from this app's own data;
# none of the template's own formulas for these fields are read or kept.
# =====================================================================

SF2_TEMPLATE_PATH = Path(__file__).resolve().parent / 'xlsx_templates' / 'SF2.xlsx'

# The 30 day-slot columns (5 weeks x Mon-Sat), left to right, as laid out in
# the template's row 17/18 header and each student row's mark cells. Only
# the top-left cell of each (possibly merged) day-slot is listed here - that
# is the one real, writable cell per slot.
DAY_COLUMNS = [
    'K', 'M', 'N', 'O', 'S', 'V',
    'W', 'Z', 'AA', 'AB', 'AE', 'AH',
    'AI', 'AL', 'AN', 'AP', 'AR', 'AS',
    'AU', 'AV', 'AW', 'AY', 'BB', 'BD',
    'BE', 'BI', 'BJ', 'BK', 'BM', 'BO',
]
MAX_DAY_SLOTS = len(DAY_COLUMNS)  # 30
MAX_STUDENTS_PER_GENDER = 55
MALE_ROW_START = 19    # rows 19-73
FEMALE_ROW_START = 75  # rows 75-129
MALE_SUBTOTAL_ROW = 74     # '<=== MALE | TOTAL Per Day ===>'
FEMALE_SUBTOTAL_ROW = 130  # '<=== FEMALE | TOTAL Per Day ===>'
COMBINED_TOTAL_ROW = 131   # 'Combined TOTAL Per Day'

# date.weekday(): Monday=0 ... Sunday=6. Normally only Mon-Sat show up here
# (Sundays are never school days by default), but a SchoolCalendarException
# can mark any date - including a Sunday - as a school day, so all 7 are covered.
WEEKDAY_ABBR = ['M', 'T', 'W', 'TH', 'F', 'S', 'SU']

# Blank = Present (never written). "L"/"C" are plain-letter stand-ins for
# the template's original diagonal half-shaded Late-Comer/Cutting-Classes
# cell styling - the one intentional visual deviation from the source file.
MARK_SYMBOL = {
    AttendanceMark.STATUS_ABSENT: 'x',
    AttendanceMark.STATUS_LATE_COMER: 'L',
    AttendanceMark.STATUS_CUTTING_CLASSES: 'C',
}

# Spelled vertically down a holiday column's student rows, starting at
# MALE_ROW_START, in place of attendance marks - one letter per row, only as
# many as fit in that gender block's visible (non-hidden) rows.
HOLIDAY_LETTERS = ('H', 'O', 'L', 'I', 'D', 'A', 'Y')

# This app has no explicit semester field (grades use Term 1/2/3, not
# semesters), so the semester shown on SF2 is inferred from where the month
# falls in ATTENDANCE_MONTHS (Jun-Oct / Nov-Apr split).
FIRST_SEMESTER_MONTHS = {6, 7, 8, 9, 10}


def _semester_for_month(month):
    return 'First Semester' if month in FIRST_SEMESTER_MONTHS else 'Second Semester'


def _sf2_student_name(student):
    parts = [f'{student.surname.strip()}, {student.name.strip()}']
    if student.middle_name:
        parts.append(student.middle_name.strip())
    return ' '.join(parts)


def _sf2_day_slot_map(school_days, holiday_days, year, month, warnings, reanchor_to_first_monday=False):
    """Maps each real school day AND each declared holiday to its
    weekday-anchored (week, weekday) slot in the 30-slot (5 weeks x Mon-Sat)
    SF2 day grid, keyed by the DAY_COLUMNS letter for that slot - not by
    list position. Monday is always a week's first slot, Saturday its
    sixth; a week with no Saturday class, a gap, or a partial first week
    (the month not starting on a Monday) all simply leave that slot out of
    the returned map rather than pulling in whatever day comes next.

    Returns (day_slot_map, holiday_columns): day_slot_map covers both
    school days and holidays (so row 17/18's date/weekday header renders
    identically for either), while holiday_columns is the subset of its
    keys that are holidays - callers use it to skip holiday columns for
    attendance-mark purposes without needing a second, undifferentiated map.

    Normally grid_start is the Monday on/before the 1st, so a month that
    doesn't start on a Monday just renders a partly-blank leading week.
    reanchor_to_first_monday skips that leading week entirely and starts
    the grid at the first Monday on/after the 1st instead - only meant for
    the Saturday/Sunday-start edge case, where the default anchor would
    otherwise waste an entire week of columns on nothing.
    """
    month_start = date(year, month, 1)
    if reanchor_to_first_monday:
        days_to_first_monday = (7 - month_start.weekday()) % 7
        grid_start = month_start + timedelta(days=days_to_first_monday)
    else:
        grid_start = month_start - timedelta(days=month_start.weekday())

    slot_map = {}
    holiday_columns = set()
    all_days = [(day, False) for day in school_days] + [(day, True) for day in holiday_days]
    for day, is_holiday in all_days:
        label = 'holiday' if is_holiday else 'school day'
        weekday = day.weekday()  # Monday=0 ... Sunday=6
        if weekday == 6:
            warnings.append(
                f"{day.isoformat()} is a {label} that falls on a Sunday; SF2's day grid "
                f"only has Monday-Saturday columns, so it was left off this form."
            )
            continue
        if reanchor_to_first_monday and day < grid_start:
            warnings.append(
                f"{day.isoformat()} is a {label} before {MONTH_NAMES[month - 1]}'s first "
                f"Monday; SF2's day grid starts there for this month, so it was left off this form."
            )
            continue
        week_number = (day - grid_start).days // 7
        slot_index = week_number * 6 + weekday
        if slot_index >= MAX_DAY_SLOTS:
            warnings.append(
                f"{day.isoformat()} falls past SF2's 5-week day grid and was left off this form."
            )
            continue
        col = DAY_COLUMNS[slot_index]
        slot_map[col] = day
        if is_holiday:
            holiday_columns.add(col)
    return slot_map, holiday_columns


def _sf2_ratio(numerator, denominator):
    return round(numerator / denominator, 4) if denominator else 0


def _fill_sf2_header(ws, teacher, section, school_profile, year, month):
    ws['I5'] = school_profile.school_name if school_profile else ''
    ws['Y5'] = school_profile.school_id if school_profile else ''
    ws['AQ5'] = school_profile.district if school_profile else ''
    ws['BH5'] = school_profile.division if school_profile else ''
    ws['BY5'] = school_profile.region if school_profile else ''
    ws['U8'] = school_profile.school_year if school_profile else ''
    ws['I7'] = _semester_for_month(month)
    ws['AN8'] = f'Grade {section.grade_level}'
    ws['I12'] = section.section_name
    ws['BC7'] = '/'.join(part for part in (section.track, section.strand) if part)
    ws['BN11'] = MONTH_NAMES[month - 1].upper()
    ws['BR168'] = teacher.full_name
    ws['BR171'] = school_profile.principal_name if school_profile else ''


def _fill_sf2_gender_block(ws, students, row_start, day_slot_map, holiday_columns, warnings, gender_label):
    """Writes one gender's student rows - only the real students, no blank
    placeholder rows - and returns the per-day/per-block stats needed for
    that gender's subtotal row and the summary section below.

    Holiday columns (holiday_columns) are skipped entirely here: no
    attendance mark is written for them, and they don't count toward a
    student's absent/present totals - a holiday isn't an instructional day,
    so it shouldn't affect ADA/attendance-% math either. They're marked
    separately afterward, once, with "HOLIDAY" spelled down the rows."""
    if len(students) > MAX_STUDENTS_PER_GENDER:
        warnings.append(
            f'{len(students)} {gender_label} students found, but SF2 only has room for '
            f'{MAX_STUDENTS_PER_GENDER}. Only the first {MAX_STUDENTS_PER_GENDER} are on this form.'
        )
        students = students[:MAX_STUDENTS_PER_GENDER]

    school_cols = [col for col in day_slot_map if col not in holiday_columns]

    lrns = [s.lrn for s in students]
    school_dates = [day_slot_map[col] for col in school_cols]
    marks_by_key = {
        f'{m.lrn}|{m.date.isoformat()}': m.status
        for m in AttendanceMark.objects.filter(lrn__in=lrns, date__in=school_dates)
    }

    present_per_col = {col: 0 for col in school_cols}
    total_absent = 0
    total_present = 0

    for i, student in enumerate(students):
        row = row_start + i
        ws[f'G{row}'] = _sf2_student_name(student)

        absent_count = 0
        for col in school_cols:
            day = day_slot_map[col]
            status = marks_by_key.get(f'{student.lrn}|{day.isoformat()}', 'present')
            if status == AttendanceMark.STATUS_ABSENT:
                absent_count += 1
            else:
                present_per_col[col] += 1
            ws[f'{col}{row}'] = MARK_SYMBOL.get(status, '')

        present_count = len(school_cols) - absent_count
        ws[f'BS{row}'] = absent_count
        ws[f'BV{row}'] = present_count
        total_absent += absent_count
        total_present += present_count

    return {
        'count': len(students),
        'present_per_col': present_per_col,
        'total_absent': total_absent,
        'total_present': total_present,
    }


def _fill_sf2_subtotal_row(ws, row, present_per_col, total_absent, total_present):
    """Writes one 'TOTAL Per Day' row (male/female/combined): per-day
    present-student counts in the day columns - matching what row 74/130's
    own (broken) template formulas compute: registered minus that day's
    absences - plus the month's ABSENT/PRESENT totals in BS/BV. Columns
    absent from present_per_col (not a school day at all, or a holiday
    column) are left blank rather than 0."""
    for col in DAY_COLUMNS:
        ws[f'{col}{row}'] = present_per_col[col] if col in present_per_col else None
    ws[f'BS{row}'] = total_absent
    ws[f'BV{row}'] = total_present


def _build_sf2_workbook(teacher, section, school_profile, year, month, reanchor_to_first_monday=False):
    """Returns (workbook, warnings). warnings is a list of human-readable
    strings for anything this export had to cap/truncate rather than
    silently drop."""
    wb = openpyxl.load_workbook(SF2_TEMPLATE_PATH)
    ws = wb['Sheet1']
    warnings = []

    # The template's own external-workbook link definitions (the source of
    # every broken [1]... formula) - drop them outright rather than just
    # overwriting the formulas that referenced them, so no dangling link
    # survives in the saved file for Excel to prompt about on open.
    wb._external_links = []

    school_days = _school_days_in_month(section.school_profile_id, year, month)
    holiday_days = list(SchoolCalendarException.objects.filter(
        school_profile_id=section.school_profile_id, date__year=year, date__month=month, is_school_day=False,
    ).values_list('date', flat=True))
    day_slot_map, holiday_columns = _sf2_day_slot_map(
        school_days, holiday_days, year, month, warnings, reanchor_to_first_monday
    )
    # Holidays occupy a column on the grid but aren't instructional days, so
    # they're excluded from the "Number of School Days" denominator below.
    school_day_count = len(day_slot_map) - len(holiday_columns)

    _fill_sf2_header(ws, teacher, section, school_profile, year, month)

    for col in DAY_COLUMNS:
        day = day_slot_map.get(col)
        ws[f'{col}17'] = day.day if day else None
        ws[f'{col}18'] = WEEKDAY_ABBR[day.weekday()] if day else None

    students = Student.objects.filter(adviser_id=teacher.teacher_id).order_by('surname', 'name')
    males = [s for s in students if s.sex in ('MALE', 'M')]
    females = [s for s in students if s.sex in ('FEMALE', 'F')]

    male_stats = _fill_sf2_gender_block(ws, males, MALE_ROW_START, day_slot_map, holiday_columns, warnings, 'male')
    female_stats = _fill_sf2_gender_block(ws, females, FEMALE_ROW_START, day_slot_map, holiday_columns, warnings, 'female')

    _fill_sf2_subtotal_row(
        ws, MALE_SUBTOTAL_ROW,
        male_stats['present_per_col'], male_stats['total_absent'], male_stats['total_present'],
    )
    _fill_sf2_subtotal_row(
        ws, FEMALE_SUBTOTAL_ROW,
        female_stats['present_per_col'], female_stats['total_absent'], female_stats['total_present'],
    )
    combined_present_per_col = {
        col: male_stats['present_per_col'][col] + female_stats['present_per_col'][col]
        for col in male_stats['present_per_col']
    }
    combined_total_absent = male_stats['total_absent'] + female_stats['total_absent']
    combined_total_present = male_stats['total_present'] + female_stats['total_present']
    _fill_sf2_subtotal_row(
        ws, COMBINED_TOTAL_ROW,
        combined_present_per_col, combined_total_absent, combined_total_present,
    )

    male_count = male_stats['count']
    female_count = female_stats['count']
    total_count = male_count + female_count

    # BZ134: "Number of School Days in reporting month" (the ADA denominator
    # documented at row 143) - uses the days actually rendered on the
    # 30-slot grid, not every real school day, so it always matches the
    # per-day totals written above (a school day that overflowed the grid
    # has no mark data on this sheet at all, so it can't be counted here
    # either - see the warnings _sf2_day_slot_map raises for that case).
    # Holiday columns are excluded too, per school_day_count above.
    ws['BZ134'] = school_day_count

    # "HOLIDAY" spelled vertically down each holiday column's student rows,
    # in place of attendance marks - done once here (not inside either
    # gender block above) since it's independent of student data. Only the
    # male block's visible rows (19-25) get the letters; a section with
    # fewer than 7 visible male rows just gets however many letters fit.
    for col in holiday_columns:
        visible_rows = min(male_count, len(HOLIDAY_LETTERS))
        for i in range(visible_rows):
            ws[f'{col}{MALE_ROW_START + i}'] = HOLIDAY_LETTERS[i]

    # "Summary" block (rows 134-148): registered-learner counts and the
    # Enrolment%/ADA/Attendance% math documented in the template itself
    # around F139-L145, replicated here in Python since the template's own
    # CC/CD/CE formulas are broken (external-workbook links, and ranges that
    # no longer line up once the unused rows below are deleted).
    ws['CC135'] = male_count
    ws['CD135'] = female_count
    ws['CE135'] = total_count

    # Late enrollment: this app has no data on when a student enrolled
    # relative to the 1st-Friday cutoff, so this stays 0 rather than guessed.
    ws['CC137'] = 0
    ws['CD137'] = 0
    ws['CE137'] = 0

    # "Registered Learners as of end of the month" - this app tracks one
    # live roster, not a point-in-time snapshot, so it's the same count.
    ws['CC142'] = male_count
    ws['CD142'] = female_count
    ws['CE142'] = total_count

    # a. Percentage of Enrolment = Registered as of end of month / Registered
    # as of 1st Friday. Both sides are the same roster above, so this is
    # always ~1 (100%) until this app tracks separate enrollment-date
    # snapshots - not fabricated, just a consequence of having one roster.
    ws['CC144'] = _sf2_ratio(male_count, male_count)
    ws['CD144'] = _sf2_ratio(female_count, female_count)
    ws['CE144'] = _sf2_ratio(total_count, total_count)

    # b. Average Daily Attendance = Total Daily Attendance / Number of
    # School Days in the reporting month.
    male_ada = _sf2_ratio(male_stats['total_present'], school_day_count)
    female_ada = _sf2_ratio(female_stats['total_present'], school_day_count)
    ws['CC146'] = male_ada
    ws['CD146'] = female_ada
    ws['CE146'] = round(male_ada + female_ada, 4)

    # c. Percentage of Attendance for the month = Average Daily Attendance /
    # Registered Learners as of end of the month.
    ws['CC148'] = _sf2_ratio(male_stats['total_present'], school_day_count * male_count)
    ws['CD148'] = _sf2_ratio(female_stats['total_present'], school_day_count * female_count)
    ws['CE148'] = _sf2_ratio(combined_total_present, school_day_count * total_count)

    # Rows 149-161 (students absent 5+ consecutive days, NLPA reason
    # breakdowns, transferred/shifted in-out) - this app has no data source
    # for any of these, so they're deliberately left at the template's own
    # static 0s rather than computed or guessed at.

    # Hide unused student rows (rather than leaving them blank) for both
    # gender blocks. This template has ~2,700 merged cell ranges, and
    # ws.delete_rows() does not reliably preserve values in merges elsewhere
    # on the sheet once saved - it silently wiped out the Summary block's
    # text (rows 132+: "Month :", "Summary", the M/F/TOTAL headers, the
    # Enrolment/ADA/Attendance% labels, etc.) on every export. Hiding
    # instead of deleting leaves every row number and merge untouched -
    # Excel simply skips hidden rows when printing - so nothing below can
    # be corrupted.
    for row in range(MALE_ROW_START + male_count, MALE_ROW_START + MAX_STUDENTS_PER_GENDER):
        ws.row_dimensions[row].hidden = True

    for row in range(FEMALE_ROW_START + female_count, FEMALE_ROW_START + MAX_STUDENTS_PER_GENDER):
        ws.row_dimensions[row].hidden = True

    return wb, warnings


@login_required
def export_sf2(request, year, month):
    try:
        teacher = _get_teacher(request)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
    if not section:
        messages.error(request, 'Your profile is missing a section. Please complete your profile first.')
        return redirect('attendance_grid')

    month_start = date(year, month, 1)
    starts_on_weekend = month_start.weekday() in (5, 6)
    if starts_on_weekend and request.GET.get('confirmed') != '1':
        weekday_name = 'Saturday' if month_start.weekday() == 5 else 'Sunday'
        confirm_url = f"{reverse('export_sf2', args=[year, month])}?confirmed=1"

        school_days = _school_days_in_month(section.school_profile_id, year, month)
        preview_slot_map, _ = _sf2_day_slot_map(school_days, [], year, month, [], reanchor_to_first_monday=True)
        first_week_cols = [col for col in DAY_COLUMNS[:6] if col in preview_slot_map]
        first_week_days = [preview_slot_map[col] for col in first_week_cols]

        if first_week_days:
            first_day, last_day = first_week_days[0], first_week_days[-1]
            date_range = (
                f"{first_day.strftime('%b')} {first_day.day}"
                if first_day == last_day
                else f"{first_day.strftime('%b')} {first_day.day}-{last_day.day}"
            )
            detail = (
                f"so this month's first school week ({date_range}) will start at column "
                f"{first_week_cols[0]} instead of its usual calendar position"
            )
        else:
            detail = "so the day grid's first calendar week has no weekdays in it"

        messages.warning(
            request,
            format_html(
                '{} {} starts on a {}, {}. <a href="{}">Generate anyway</a>',
                MONTH_NAMES[month - 1], year, weekday_name, detail, confirm_url,
            ),
        )
        return redirect(f"{reverse('attendance_grid')}?year={year}&month={month}")

    school_profile = None
    if teacher.school_profile_id:
        school_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()

    wb, warnings = _build_sf2_workbook(
        teacher, section, school_profile, year, month, reanchor_to_first_monday=starts_on_weekend
    )
    for warning in warnings:
        messages.warning(request, warning)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'SF2_{section.section_name}_{MONTH_NAMES[month - 1]}{year}.xlsx'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
