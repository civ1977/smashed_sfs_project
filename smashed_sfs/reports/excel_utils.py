from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

THIN = Side(style='thin', color='999999')
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def merge_and_write(ws, row, col_start, col_end, text, bold=False, size=11, align='center'):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align)
    return cell


def write_table_header(ws, row, col_start, headers):
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center')


def write_table_row(ws, row, col_start, values):
    for i, value in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=value)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center')


def excel_response(workbook, filename):
    from django.http import HttpResponse

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
