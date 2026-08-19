import openpyxl
from openpyxl.styles import Font, PatternFill

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from .models import Student, Section, log_student_record_change
from accounts.models import Teacher
from portal.models import StudentAccount
from smashed_sfs.upload_utils import read_upload_rows, UploadFileError

STUDENT_TEMPLATE_HEADERS = [
    'LRN', 'Surname', 'Name', 'Middle Name', 'Extension', 'Sex', 'Birthday',
    'School (G10)', 'School Address', 'Average', 'Completion Date', 'SHS Admission',
]
STUDENT_TEMPLATE_SAMPLE_ROW = [
    '123456789012', 'Dela Cruz', 'Juan', 'Santos', '', 'MALE', '2008-06-30',
    'Sample National High School', 'Sample Address, Sample City', '90.5', '2024-06-05', '2024-08-05',
]


def convert_date(date_str):
    """Convert various date formats to YYYY-MM-DD"""
    if not date_str or date_str.strip() == '':
        return None
    
    date_str = date_str.strip()
    
    formats = [
        '%b %d, %Y',    # JUN 30, 2008
        '%B %d, %Y',    # June 30, 2008
        '%m/%d/%Y',     # 06/30/2008
        '%m-%d-%Y',     # 06-30-2008
        '%Y-%m-%d',     # 2008-06-30
        '%Y/%m/%d',     # 2008/06/30
        '%d-%b-%Y',     # 30-JUN-2008
        '%d %b %Y',     # 30 JUN 2008
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    
    return date_str


@login_required
def upload_students(request):
    preview_data = None
    
    if request.method == 'POST' and 'csv_file' in request.FILES:
        csv_file = request.FILES['csv_file']
        skip_header = request.POST.get('skip_header') == 'on'
        
        try:
            rows = iter(read_upload_rows(csv_file))

            if skip_header:
                next(rows, None)

            preview_data = [row for row in rows if row]

            request.session['student_preview_data'] = preview_data

            messages.info(request, f'📋 Preview loaded: {len(preview_data)} students found.')

        except UploadFileError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error reading file: {str(e)}')
    
    if 'student_preview_data' in request.session and not preview_data:
        preview_data = request.session.get('student_preview_data')
    
    return render(request, 'students/upload.html', {'preview_data': preview_data})


@login_required
def save_students(request):
    if request.method == 'POST':
        row_count = int(request.POST.get('row_count', 0))
        
        saved_count = 0
        error_count = 0
        errors = []
        
        # Get teacher by username
        try:
            teacher = Teacher.objects.get(username=request.user.username)
        except Teacher.DoesNotExist:
            messages.error(request, 'Your account is not linked to a Teacher profile. Please contact the administrator.')
            return redirect('upload_students')
        
        # Get the teacher's section
        section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
        
        if not section:
            messages.error(request, 'No section found for this teacher. Please complete your profile first.')
            return redirect('complete_profile')
        
        for i in range(row_count):
            try:
                lrn = request.POST.get(f'lrn_{i}', '').strip()
                if not lrn:
                    continue
                
                # Check if student already exists
                if Student.objects.filter(lrn=lrn).exists():
                    messages.warning(request, f'Student {lrn} already exists. Skipping.')
                    continue
                
                # Get date fields and convert them
                birthday = convert_date(request.POST.get(f'birthday_{i}', ''))
                completion_date = convert_date(request.POST.get(f'completion_date_g10_{i}', ''))
                shs_admission = convert_date(request.POST.get(f'shs_admission_date_{i}', ''))
                
                # Get average, handle empty value
                avg_str = request.POST.get(f'average_g10_{i}', '0').strip()
                try:
                    avg = float(avg_str) if avg_str else 0
                except ValueError:
                    avg = 0
                
                student = Student(
                    lrn=lrn,
                    surname=request.POST.get(f'surname_{i}', '').strip(),
                    name=request.POST.get(f'name_{i}', '').strip(),
                    middle_name=request.POST.get(f'middle_name_{i}', '').strip() or None,
                    extension=request.POST.get(f'extension_{i}', '').strip() or None,
                    sex=request.POST.get(f'sex_{i}', 'MALE'),
                    birthday=birthday,
                    school_g10=request.POST.get(f'school_g10_{i}', '').strip(),
                    school_address_g10=request.POST.get(f'school_address_g10_{i}', '').strip(),
                    average_g10=avg,
                    completion_date_g10=completion_date,
                    shs_admission_date=shs_admission,
                    section_id=section.section_id,
                    adviser_id=teacher.teacher_id
                )
                student.save()
                log_student_record_change('student', student.lrn, student.lrn, 'created', teacher.teacher_id)
                saved_count += 1
                
            except Exception:
                error_count += 1
                error_msg = f"Row {i+1} couldn't be saved - please check its values and try again."
                errors.append(error_msg)
                messages.error(request, error_msg)
        
        request.session.pop('student_preview_data', None)
        
        if saved_count > 0:
            messages.success(request, f'✅ {saved_count} students saved successfully!')
        if error_count > 0:
            messages.warning(request, f'{error_count} students had errors.')
        
        return redirect('student_list')
    
    return redirect('upload_students')


@login_required
def download_student_template(request):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Students'

    sheet.append(STUDENT_TEMPLATE_HEADERS)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    sheet.append(STUDENT_TEMPLATE_SAMPLE_ROW)

    widths = [14, 14, 14, 14, 10, 10, 12, 24, 24, 10, 16, 16]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_upload_template.xlsx"'
    wb.save(response)
    return response


@login_required
def student_list(request):
    try:
        teacher = Teacher.objects.get(username=request.user.username)
        ordered_students = Student.objects.filter(adviser_id=teacher.teacher_id).order_by('surname', 'name')
        males = [s for s in ordered_students if s.sex in ('MALE', 'M')]
        females = [s for s in ordered_students if s.sex in ('FEMALE', 'F')]
        students = males + females
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        students = []

    return render(request, 'students/list.html', {'students': students})


@login_required
def update_student(request, lrn):
    if request.method != 'POST':
        return redirect('student_list')

    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('student_list')

    student = Student.objects.filter(lrn=lrn, adviser_id=teacher.teacher_id).first()
    if not student:
        messages.error(request, 'Student not found.')
        return redirect('student_list')

    student.surname = request.POST.get('surname', '').strip()
    student.name = request.POST.get('name', '').strip()
    student.middle_name = request.POST.get('middle_name', '').strip() or None
    student.sex = request.POST.get('sex', student.sex)
    student.birthday = convert_date(request.POST.get('birthday', ''))
    status = request.POST.get('status', student.status)
    if status in dict(Student.STATUS_CHOICES):
        student.status = status
    student.save()
    log_student_record_change('student', student.lrn, student.lrn, 'updated', teacher.teacher_id)

    messages.success(request, f'✅ Updated {student.surname}, {student.name}.')
    return redirect('student_list')


@login_required
def access_requests(request):
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    my_lrns = Student.objects.filter(adviser_id=teacher.teacher_id).values_list('lrn', flat=True)
    pending_accounts = StudentAccount.objects.filter(
        status=StudentAccount.STATUS_PENDING,
        lrn__in=my_lrns
    ).order_by('requested_at')

    students_by_lrn = {
        s.lrn: s for s in Student.objects.filter(lrn__in=[a.lrn for a in pending_accounts])
    }
    requests = [
        {'account': account, 'student': students_by_lrn.get(account.lrn)}
        for account in pending_accounts
    ]

    return render(request, 'students/access_requests.html', {'requests': requests})


@login_required
def decide_access_request(request, account_id):
    if request.method != 'POST':
        return redirect('access_requests')

    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('access_requests')

    account = get_object_or_404(StudentAccount, account_id=account_id, status=StudentAccount.STATUS_PENDING)

    # Only this adviser's own students - never another section's requests.
    student = Student.objects.filter(lrn=account.lrn, adviser_id=teacher.teacher_id).first()
    if not student:
        messages.error(request, 'That request does not belong to one of your students.')
        return redirect('access_requests')

    decision = request.POST.get('decision')
    if decision == 'approve':
        account.status = StudentAccount.STATUS_APPROVED
        messages.success(request, f'✅ Approved account access for {student.surname}, {student.name}.')
    elif decision == 'reject':
        account.status = StudentAccount.STATUS_REJECTED
        messages.info(request, f'Rejected account access for {student.surname}, {student.name}.')
    else:
        messages.error(request, 'Invalid decision.')
        return redirect('access_requests')

    account.decided_at = timezone.now()
    account.decided_by = teacher.teacher_id
    account.save()

    return redirect('access_requests')