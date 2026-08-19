import openpyxl
from openpyxl.styles import Font, PatternFill

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.utils.dateformat import format as format_date
import calendar
import json
import statistics
from datetime import datetime, date as date_cls
from decimal import Decimal, ROUND_HALF_UP
from students.models import Student, Section, SchoolProfile, log_student_record_change
from students.utils import male_then_female
from accounts.models import Teacher
from smashed_sfs.upload_utils import read_upload_rows, UploadFileError
from .models import Grade, SubjectMapping, SubjectTermExclusion, StudentTermRemark, Attendance, ATTENDANCE_MONTHS, AttendanceMark, SchoolCalendarException, TeacherSubjectAssignment, SubjectTestMaxScore, AncillaryTask, AncillaryTaskSchedule


def round_half_up(value, ndigits=0):
    """Round a grade average the way DepEd does: halves always round up
    (88.5 -> 89), not Python's built-in round-half-to-even (88.5 -> 88,
    89.5 -> 90). Only for official grade averages - not for statistics
    like mean/stdev/MPS or attendance ratios, which aren't transmuted
    grades and can keep normal rounding.
    """
    quantum = Decimal(1).scaleb(-ndigits)
    quantized = Decimal(str(value)).quantize(quantum, rounding=ROUND_HALF_UP)
    return int(quantized) if ndigits == 0 else float(quantized)


MONTH_NAMES = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December',
]

GRADE_TERMS = [
    {'value': 1, 'label': '1st Term'},
    {'value': 2, 'label': '2nd Term'},
    {'value': 3, 'label': '3rd Term'},
]

ATTENDANCE_MONTH_OPTIONS = [{'value': m, 'label': m} for m in ATTENDANCE_MONTHS]

GRADE_LEVELS = [
    {'value': str(g), 'label': f'Grade {g}'} for g in range(4, 13)
]


def convert_date(date_str):
    """Convert various date formats to YYYY-MM-DD"""
    if not date_str or date_str.strip() == '':
        return None
    
    date_str = date_str.strip()
    
    formats = [
        '%b %d, %Y',
        '%B %d, %Y',
        '%m/%d/%Y',
        '%m-%d-%Y',
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%d-%b-%Y',
        '%d %b %Y',
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    
    return date_str


@login_required
def upload_grades(request):
    preview_data = None
    term = request.POST.get('term') if request.method == 'POST' else None
    grade_level = request.POST.get('grade_level') if request.method == 'POST' else None
    headers = None

    if request.method == 'POST' and 'csv_file' in request.FILES:
        csv_file = request.FILES['csv_file']
        term = request.POST.get('term')
        grade_level = request.POST.get('grade_level')
        skip_header = request.POST.get('skip_header') == 'on'
        replace_all = request.POST.get('replace_all') == 'on'

        try:
            rows = iter(read_upload_rows(csv_file))

            preview_data = []
            headers = None

            if skip_header:
                try:
                    headers = next(rows)
                    headers = [h.strip() if h else f'Subject{i+1}' for i, h in enumerate(headers)]
                except StopIteration:
                    pass

            for row in rows:
                if row:
                    if headers and len(row) < len(headers):
                        row.extend([''] * (len(headers) - len(row)))
                    preview_data.append(row)

            if not headers and preview_data:
                headers = ['LRN'] + [f'Subject{i+1}' for i in range(len(preview_data[0]) - 1)]

            request.session['grade_preview_data'] = preview_data
            request.session['grade_term'] = term
            request.session['grade_level'] = grade_level
            request.session['grade_replace_all'] = replace_all
            request.session['grade_headers'] = headers

            messages.info(request, f'📋 Preview loaded: {len(preview_data)} students found for Term {term}.')

        except UploadFileError as e:
            messages.error(request, str(e))
        except Exception:
            messages.error(request, "That file couldn't be read — check that it's a valid CSV/Excel export and try again.")

    if 'grade_preview_data' in request.session and not preview_data:
        preview_data = request.session.get('grade_preview_data')
        term = request.session.get('grade_term')
        grade_level = request.session.get('grade_level')
        headers = request.session.get('grade_headers')

    return render(request, 'grades/upload.html', {
        'preview_data': preview_data,
        'term': term,
        'terms': GRADE_TERMS,
        'grade_level': grade_level or '11',
        'grade_levels': GRADE_LEVELS,
        'headers': headers,
    })


@login_required
def save_grades(request):
    if request.method == 'POST':
        row_count = int(request.POST.get('row_count', 0))
        term = int(request.POST.get('term', 1))
        grade_level = request.POST.get('grade_level') or '11'
        replace_all = request.POST.get('replace_all') == 'on'
        update_subjects = request.POST.get('update_subjects') == 'on'
        
        saved_count = 0
        error_count = 0
        errors = []
        
        # Get teacher
        try:
            teacher = Teacher.objects.get(username=request.user.username)
        except Teacher.DoesNotExist:
            messages.error(request, 'Your account is not linked to a Teacher profile.')
            return redirect('upload_grades')
        
        if not teacher.school_profile_id:
            messages.error(request, 'Your profile is missing a school. Please complete your profile first.')
            return redirect('complete_profile')

        # Subjects are scoped by the adviser's own section strand (matching
        # arrange_subjects/school.adviser_subject_assignments), not a
        # hardcoded strand - a blank section.strand means the section makes
        # no strand distinction, so mappings are created/looked up as blank too.
        teacher_section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
        strand = teacher_section.strand if teacher_section else ''

        # Get subject mapping
        subject_mappings = {}
        for i in range(1, 31):
            subject_name = request.POST.get(f'subject_name_{i}')
            if subject_name and update_subjects:
                mapping, created = SubjectMapping.objects.get_or_create(
                    school_profile_id=teacher.school_profile_id,
                    grade_level=grade_level,
                    strand=strand,
                    subject_number=i,
                    defaults={
                        'subject_name': subject_name,
                        'created_by': teacher.teacher_id
                    }
                )
                if not created and mapping.subject_name != subject_name:
                    mapping.subject_name = subject_name
                    mapping.save()
                subject_mappings[i] = mapping.mapping_id
            else:
                try:
                    mapping = SubjectMapping.objects.get(
                        school_profile_id=teacher.school_profile_id,
                        grade_level=grade_level,
                        strand=strand,
                        subject_number=i
                    )
                    subject_mappings[i] = mapping.mapping_id
                except SubjectMapping.DoesNotExist:
                    pass
        
        # Delete existing grades for this term if replace_all - scoped to this
        # grade level's mappings only, so a Grade 12 replace doesn't wipe a
        # student's Grade 11 history for the same term number.
        if replace_all:
            grade_level_mapping_ids = set(subject_mappings.values())
            students = Student.objects.filter(adviser_id=teacher.teacher_id)
            for student in students:
                to_delete = Grade.objects.filter(
                    lrn=student.lrn,
                    term=term,
                    mapping_id__in=grade_level_mapping_ids
                )
                for grade in to_delete:
                    log_student_record_change('grade', grade.grade_id, grade.lrn, 'deleted', teacher.teacher_id)
                to_delete.delete()
            messages.info(request, f'🗑️ Existing Grade {grade_level} Term {term} grades deleted.')
        
        for i in range(row_count):
            try:
                lrn = request.POST.get(f'lrn_{i}', '').strip()
                if not lrn:
                    continue

                try:
                    student = Student.objects.get(lrn=lrn, adviser_id=teacher.teacher_id)
                except Student.DoesNotExist:
                    messages.warning(request, f'Student {lrn} not found. Skipping.')
                    continue

                # A savepoint per row (not one for the whole loop) - a
                # DB-level error partway through this row's up-to-30
                # subjects shouldn't leave it half-saved, but one bad row
                # also shouldn't sour the transaction for every row after
                # it, which the existing per-row try/except already
                # tolerates (see error_count/saved_count reporting below).
                with transaction.atomic():
                    for subject_num in range(1, 31):
                        grade_value = request.POST.get(f'grade_{i}_{subject_num}', '').strip()
                        if grade_value:
                            try:
                                grade = int(grade_value)
                                if grade < 60 or grade > 100:
                                    messages.warning(request, f'Grade {grade} for {lrn} is outside 60-100 range.')
                                    continue

                                mapping_id = subject_mappings.get(subject_num)
                                if mapping_id:
                                    existing = Grade.objects.filter(
                                        lrn=lrn,
                                        mapping_id=mapping_id,
                                        term=term
                                    ).first()

                                    if existing:
                                        existing.grade = grade
                                        existing.save()
                                        log_student_record_change('grade', existing.grade_id, lrn, 'updated', teacher.teacher_id)
                                    else:
                                        new_grade = Grade.objects.create(
                                            lrn=lrn,
                                            mapping_id=mapping_id,
                                            term=term,
                                            grade=grade,
                                            uploaded_by=teacher.teacher_id
                                        )
                                        log_student_record_change('grade', new_grade.grade_id, lrn, 'created', teacher.teacher_id)
                                    saved_count += 1
                            except ValueError:
                                messages.warning(request, f'Invalid grade value for {lrn}: {grade_value}')

            except Exception:
                error_count += 1
                row_error = f"Row {i+1} couldn't be saved - please check its values and try again."
                errors.append(row_error)
                messages.error(request, row_error)
        
        request.session.pop('grade_preview_data', None)
        request.session.pop('grade_term', None)
        request.session.pop('grade_level', None)
        request.session.pop('grade_headers', None)

        if saved_count > 0:
            messages.success(request, f'✅ {saved_count} grades saved successfully for Term {term}!')
        if error_count > 0:
            messages.warning(request, f'{error_count} students had errors.')
        
        return redirect('view_grades', lrn='all')
    
    return redirect('upload_grades')


def build_student_grade_sheet(lrn):
    """Shared gradesheet data-shaping for a single student: per-subject
    term grades, final average, and remarks, scoped to the student's own
    current grade level via their Section. Used by the adviser-facing
    view_grades(lrn) below and by the student portal (portal.views), so
    the shaping logic isn't duplicated between the two.

    Returns (subject_grades, subject_names, general_average, grades).
    """
    student = Student.objects.filter(lrn=lrn).first()
    student_section = Section.objects.filter(section_id=student.section_id).first() if student else None

    grades = Grade.objects.filter(lrn=lrn).order_by('term', 'mapping_id')
    if student_section:
        grade_level_mapping_ids = set(SubjectMapping.objects.filter(
            school_profile_id=student_section.school_profile_id,
            grade_level=student_section.grade_level
        ).values_list('mapping_id', flat=True))
        grades = grades.filter(mapping_id__in=grade_level_mapping_ids)

    subject_grades = {}
    incomplete_by_subject = {}
    for grade in grades:
        if grade.mapping_id not in subject_grades:
            subject_grades[grade.mapping_id] = {1: None, 2: None, 3: None}
        subject_grades[grade.mapping_id][grade.term] = grade.grade
        if grade.is_incomplete:
            incomplete_by_subject[grade.mapping_id] = True

    all_complete = True
    for mapping_id in subject_grades:
        # A subject marked Incomplete in any term reports as INC overall,
        # regardless of whether Term 3 has been entered - it's a
        # deliberate status, not a blank waiting on data.
        if incomplete_by_subject.get(mapping_id):
            subject_grades[mapping_id]['final'] = None
            subject_grades[mapping_id]['remarks'] = 'INC'
            subject_grades[mapping_id]['is_incomplete'] = True
            all_complete = False
            continue

        subject_grades[mapping_id]['is_incomplete'] = False
        if subject_grades[mapping_id][3] is None:
            subject_grades[mapping_id]['final'] = None
            subject_grades[mapping_id]['remarks'] = None
            all_complete = False
            continue

        grades_list = [
            subject_grades[mapping_id][1],
            subject_grades[mapping_id][2],
            subject_grades[mapping_id][3]
        ]
        valid_grades = [g for g in grades_list if g is not None]
        subject_grades[mapping_id]['final'] = round_half_up(sum(valid_grades) / len(valid_grades)) if valid_grades else None
        subject_grades[mapping_id]['remarks'] = 'Passed' if subject_grades[mapping_id]['final'] and subject_grades[mapping_id]['final'] >= 75 else 'Failed'

    fetched_names = dict(
        SubjectMapping.objects.filter(mapping_id__in=subject_grades.keys())
        .values_list('mapping_id', 'subject_name')
    )
    subject_names = {
        mapping_id: fetched_names.get(mapping_id, f'Subject {mapping_id}')
        for mapping_id in subject_grades.keys()
    }

    finals = [data['final'] for data in subject_grades.values() if data['final'] is not None]
    general_average = round_half_up(sum(finals) / len(finals)) if (all_complete and finals) else None

    return subject_grades, subject_names, general_average, grades


# Assumption, not data this app has: DepEd's standard published honor-roll
# cutoffs, each also gated on no failing subject grade. The sample
# Rank.xlsm's actual awards didn't cleanly reduce to an average-only rule
# and may depend on a criterion (conduct, a specific subject floor, etc.)
# this app has no data for - surfaced as a UI note on the rankings page
# too, so confirm against the school's actual policy before relying on it.
HONOR_HIGHEST_MIN = 98
HONOR_HIGH_MIN = 95
HONOR_WITH_MIN = 90
PASSING_GRADE = 75


def _award_for_average(general_average, has_failing_grade):
    if has_failing_grade:
        return None
    if general_average >= HONOR_HIGHEST_MIN:
        return 'With Highest Honors'
    if general_average >= HONOR_HIGH_MIN:
        return 'With High Honors'
    if general_average >= HONOR_WITH_MIN:
        return 'With Honors'
    return None


def _compute_term_rankings(teacher):
    """Per-term class rankings, shared by the rankings page and its
    downloadable Excel export so both agree on scoring, ties, and
    ordering."""
    students = Student.objects.filter(adviser_id=teacher.teacher_id).order_by('surname', 'name')
    students = male_then_female(students)

    # Same grade-level scoping as view_grades('all'): a subject only
    # "counts" for completeness/ranking if it's mapped for this teacher's
    # own grade level.
    teacher_section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
    subjects = SubjectMapping.objects.filter(school_profile_id=teacher.school_profile_id)
    if teacher_section:
        subjects = subjects.filter(grade_level=teacher_section.grade_level)
    base_subject_ids = set(subjects.values_list('mapping_id', flat=True))

    # A subject removed from a specific term (SubjectTermExclusion) doesn't
    # count toward that term's ranking-completeness requirement, even though
    # it's still required for whichever other terms it isn't excluded from.
    exclusions_by_term = {1: set(), 2: set(), 3: set()}
    for exclusion in SubjectTermExclusion.objects.filter(mapping_id__in=base_subject_ids):
        exclusions_by_term.setdefault(exclusion.term, set()).add(exclusion.mapping_id)

    # Each student's raw per-term subject grades, computed once and reused
    # across all three terms below. One batched query instead of calling
    # build_student_grade_sheet per student (which itself issued several
    # queries per call) - all these students share the same section/grade
    # level (one adviser, one section), so base_subject_ids above already
    # matches what build_student_grade_sheet would resolve per student.
    student_subject_grades = {student.lrn: {} for student in students}
    for g in Grade.objects.filter(lrn__in=[s.lrn for s in students], mapping_id__in=base_subject_ids):
        subject_grades = student_subject_grades[g.lrn]
        if g.mapping_id not in subject_grades:
            subject_grades[g.mapping_id] = {1: None, 2: None, 3: None}
        subject_grades[g.mapping_id][g.term] = g.grade

    remarks_by_key = {
        (r.lrn, r.term): r.remark
        for r in StudentTermRemark.objects.filter(lrn__in=[s.lrn for s in students])
    }

    terms = []
    for term in (1, 2, 3):
        required_subject_ids = base_subject_ids - exclusions_by_term.get(term, set())

        # Terms are ranked independently off subject_grades[mapping_id][term]
        # (the raw grade entered for that term), not the cross-term 'final' -
        # a student missing Term 3 still ranks in Term 1/2 on what they have.
        complete = []
        incomplete_students = []
        for student in students:
            subject_grades = student_subject_grades[student.lrn]
            term_values = {}
            missing_subject = not required_subject_ids
            for mapping_id in required_subject_ids:
                value = subject_grades.get(mapping_id, {}).get(term)
                if value is None:
                    missing_subject = True
                    break
                term_values[mapping_id] = value

            if missing_subject:
                incomplete_students.append(student)
                continue

            general_average = round_half_up(sum(term_values.values()) / len(term_values))
            has_failing_grade = any(value < PASSING_GRADE for value in term_values.values())
            complete.append({
                'student': student,
                'general_average': general_average,
                'has_failing_grade': has_failing_grade,
                'remark': remarks_by_key.get((student.lrn, term), ''),
            })

        # Standard competition ranking (1224): ties share a rank, and the next
        # distinct average is ranked by its position, skipping the ranks the
        # tie consumed - e.g. two students tied at position 6 are both rank 6,
        # and the next student down is rank 8, not 7.
        complete.sort(key=lambda row: row['general_average'], reverse=True)
        rank = 0
        previous_average = None
        award_counts = {'With Highest Honors': 0, 'With High Honors': 0, 'With Honors': 0}
        for position, row in enumerate(complete, start=1):
            if row['general_average'] != previous_average:
                rank = position
            row['rank'] = rank
            previous_average = row['general_average']

            row['award'] = _award_for_average(row['general_average'], row['has_failing_grade'])
            if row['award']:
                award_counts[row['award']] += 1

        award_summary = [
            {'label': 'With Highest Honors', 'count': award_counts['With Highest Honors']},
            {'label': 'With High Honors', 'count': award_counts['With High Honors']},
            {'label': 'With Honors', 'count': award_counts['With Honors']},
        ]

        # Displayed male-then-female, alphabetical - same roster convention
        # used everywhere else in the app - now that each row already has
        # its rank/award computed from the descending-average sort above.
        complete.sort(key=lambda row: (row['student'].sex not in ('MALE', 'M'), row['student'].surname, row['student'].name))

        terms.append({
            'term': term,
            'rankings': complete,
            'incomplete_students': incomplete_students,
            'award_summary': award_summary,
        })

    return terms


@login_required
def view_rankings(request):
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    terms = _compute_term_rankings(teacher)

    return render(request, 'grades/rankings.html', {
        'teacher': teacher,
        'terms': terms,
    })


@login_required
def download_rankings(request):
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    terms = _compute_term_rankings(teacher)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    headers = ['No.', 'Name of Learner', 'Average', 'Rank', 'Award']

    for term_data in terms:
        sheet = wb.create_sheet(title=f'Term {term_data["term"]}')
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        for i, row in enumerate(term_data['rankings'], start=1):
            student = row['student']
            name = f'{student.surname}, {student.name}'
            if student.middle_name:
                name += f' {student.middle_name}'
            sheet.append([i, name, row['general_average'], row['rank'], row['award'] or ''])

        widths = [6, 32, 12, 8, 22]
        for col_i, width in enumerate(widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="class_rankings.xlsx"'
    wb.save(response)
    return response


def _class_students_and_subjects(teacher, term):
    """Students and subjects for this teacher's own class gradesheet -
    shared by the 'view all' gradesheet and its downloadable upload
    template, so both agree on ordering and on which subjects are
    excluded for the given term."""
    ordered_students = Student.objects.filter(adviser_id=teacher.teacher_id).order_by('surname', 'name')
    students = male_then_female(ordered_students)

    teacher_section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
    subjects = SubjectMapping.objects.filter(school_profile_id=teacher.school_profile_id)
    if teacher_section:
        subjects = subjects.filter(grade_level=teacher_section.grade_level)
    subjects = list(subjects.order_by('subject_number'))

    excluded_mapping_ids = set(
        SubjectTermExclusion.objects.filter(
            term=term, mapping_id__in=[s.mapping_id for s in subjects]
        ).values_list('mapping_id', flat=True)
    )
    excluded_subjects = [s for s in subjects if s.mapping_id in excluded_mapping_ids]
    included_subjects = [s for s in subjects if s.mapping_id not in excluded_mapping_ids]

    return students, included_subjects, excluded_subjects


@login_required
def download_grade_template(request):
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    try:
        term = int(request.GET.get('term', 1))
    except ValueError:
        term = 1
    if term not in (1, 2, 3):
        term = 1

    students, subjects, _ = _class_students_and_subjects(teacher, term)
    subject_names = [s.subject_name for s in subjects] or ['Subject1']

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f'Term {term} Grades'

    sheet.append(['LRN'] + subject_names)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for student in students:
        sheet.append([student.lrn] + [''] * len(subject_names))

    sheet.column_dimensions['A'].width = 14
    for i in range(len(subject_names)):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i + 2)].width = 14

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'grade_upload_template_term{term}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def view_grades(request, lrn):
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')
    
    if lrn == 'all':
        try:
            term = int(request.GET.get('term', 1))
        except ValueError:
            term = 1
        if term not in (1, 2, 3):
            term = 1

        students, subjects, excluded_subjects = _class_students_and_subjects(teacher, term)
        subject_ids = {s.mapping_id for s in subjects}

        grades_data = {student.lrn: {} for student in students}
        all_grades = Grade.objects.filter(lrn__in=[s.lrn for s in students], mapping_id__in=subject_ids)
        for grade in all_grades:
            subject_grades = grades_data[grade.lrn]
            if grade.mapping_id not in subject_grades:
                subject_grades[grade.mapping_id] = {1: None, 2: None, 3: None}
            subject_grades[grade.mapping_id][grade.term] = grade.grade

        return render(request, 'grades/list.html', {
            'students': students,
            'grades_data': grades_data,
            'subjects': subjects,
            'excluded_subjects': excluded_subjects,
            'teacher': teacher,
            'term': term,
            'terms': GRADE_TERMS,
        })
    else:
        student = get_object_or_404(Student, lrn=lrn, adviser_id=teacher.teacher_id)
        subject_grades, subject_names, general_average, grades = build_student_grade_sheet(lrn)

        return render(request, 'grades/view.html', {
            'student': student,
            'subject_grades': subject_grades,
            'subject_names': subject_names,
            'general_average': general_average,
            'grades': grades,
        })


def _term_from_post(request):
    try:
        term = int(request.POST.get('term', 1))
    except ValueError:
        term = 1
    return term if term in (1, 2, 3) else 1


@login_required
def arrange_subjects(request):
    """Lets an adviser drag their grade level's subjects between Core and
    Elective, and reorder within each group - saved into
    SubjectMapping.is_elective and .subject_number, the same fields
    reports/views.py's _split_core_elective and _build_subject_rows (and
    the class gradesheet's own .order_by('subject_number')) already read,
    so this one arrangement drives the gradesheet, SF9, and SF10
    identically rather than each picking its own grouping/sort."""
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    teacher_section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
    if not teacher_section:
        messages.error(request, 'No section found for this teacher. Please complete your profile first.')
        return redirect('complete_profile')

    school_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()

    # Same scoping as school/views.py's adviser_subject_assignments (the
    # subject-teacher assignment page these subjects are also managed
    # from): active subjects only, and a blank mapping.strand applies to
    # every strand at this grade level while a blank section.strand means
    # this section itself makes no strand distinction at all.
    subjects = SubjectMapping.objects.filter(
        school_profile_id=teacher.school_profile_id,
        grade_level=teacher_section.grade_level,
        is_active=True,
    )
    if teacher_section.strand:
        subjects = subjects.filter(Q(strand='') | Q(strand=teacher_section.strand))
    subjects = subjects.order_by('subject_number', 'subject_name')

    if request.method == 'POST':
        try:
            core_ids = [int(v) for v in request.POST.getlist('core_id')]
            elective_ids = [int(v) for v in request.POST.getlist('elective_id')]
        except ValueError:
            messages.error(request, 'Something went wrong reading the new order - please try again.')
            return redirect('arrange_subjects')

        ordered_ids = core_ids + elective_ids
        valid_ids = set(subjects.values_list('mapping_id', flat=True))
        if set(ordered_ids) != valid_ids or len(ordered_ids) != len(valid_ids):
            messages.error(request, 'The subject list changed before you saved - please try again.')
            return redirect('arrange_subjects')

        with transaction.atomic():
            # Two passes so a subject never briefly lands on a number
            # another subject in this same grade level still holds - safe
            # even if a future migration reintroduces a unique constraint
            # on (school_profile_id, grade_level, strand, subject_number).
            for offset, mapping_id in enumerate(ordered_ids, start=1):
                SubjectMapping.objects.filter(mapping_id=mapping_id).update(subject_number=100000 + offset)
            for position, mapping_id in enumerate(ordered_ids, start=1):
                SubjectMapping.objects.filter(mapping_id=mapping_id).update(subject_number=position)
            SubjectMapping.objects.filter(mapping_id__in=core_ids).update(is_elective=False)
            SubjectMapping.objects.filter(mapping_id__in=elective_ids).update(is_elective=True)

        messages.success(request, 'Subject arrangement saved - the gradesheet, SF9, and SF10 now show this same grouping and order.')
        return redirect('arrange_subjects')

    return render(request, 'grades/arrange_subjects.html', {
        'core_subjects': [s for s in subjects if not s.is_elective],
        'elective_subjects': [s for s in subjects if s.is_elective],
        'teacher_section': teacher_section,
        'show_core_heading': school_profile.show_core_heading if school_profile else True,
        'show_elective_heading': school_profile.show_elective_heading if school_profile else True,
    })


@login_required
def toggle_subject_group_heading(request, group):
    """Hide/unhide the 'Core Subjects'/'Elective Subjects' label row on
    SF9/SF10 (school-wide, via SchoolProfile.show_core_heading /
    .show_elective_heading) - the subjects underneath still render either
    way, only the group's own heading row is affected. Some schools don't
    distinguish Core vs Elective at all and want a plain, ungrouped list."""
    if request.method != 'POST' or group not in ('core', 'elective'):
        return redirect('arrange_subjects')

    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    school_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()
    if not school_profile:
        messages.error(request, 'No school profile found. Please complete your profile first.')
        return redirect('complete_profile')

    field_name = 'show_core_heading' if group == 'core' else 'show_elective_heading'
    new_value = not getattr(school_profile, field_name)
    setattr(school_profile, field_name, new_value)
    school_profile.save(update_fields=[field_name, 'updated_at'])

    label = 'Core Subjects' if group == 'core' else 'Elective Subjects'
    messages.success(request, f'"{label}" heading is now {"shown" if new_value else "hidden"} on SF9/SF10.')
    return redirect('arrange_subjects')


@login_required
def exclude_subject_term(request, mapping_id):
    if request.method != 'POST':
        return redirect('view_grades', lrn='all')

    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    term = _term_from_post(request)

    mapping = SubjectMapping.objects.filter(
        mapping_id=mapping_id, school_profile_id=teacher.school_profile_id
    ).first()
    if not mapping:
        messages.error(request, 'Subject not found in your school.')
    else:
        SubjectTermExclusion.objects.get_or_create(
            mapping_id=mapping.mapping_id, term=term,
            defaults={'excluded_by': teacher.teacher_id},
        )
        messages.success(request, f'✅ {mapping.subject_name} removed from Term {term}.')

    return redirect(f"{reverse('view_grades', args=['all'])}?term={term}")


@login_required
def restore_subject_term(request, mapping_id):
    if request.method != 'POST':
        return redirect('view_grades', lrn='all')

    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    term = _term_from_post(request)

    mapping = SubjectMapping.objects.filter(
        mapping_id=mapping_id, school_profile_id=teacher.school_profile_id
    ).first()
    if not mapping:
        messages.error(request, 'Subject not found in your school.')
    else:
        SubjectTermExclusion.objects.filter(mapping_id=mapping.mapping_id, term=term).delete()
        messages.success(request, f'✅ {mapping.subject_name} restored for Term {term}.')

    return redirect(f"{reverse('view_grades', args=['all'])}?term={term}")


@login_required
def save_term_remark(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        return JsonResponse({'error': 'Your account is not linked to a Teacher profile.'}, status=403)

    if teacher.role != Teacher.ROLE_ADVISER:
        return JsonResponse({'error': 'Adviser accounts only.'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (ValueError, UnicodeDecodeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    lrn = payload.get('lrn', '')
    try:
        term = int(payload.get('term'))
    except (TypeError, ValueError):
        term = None
    remark = (payload.get('remark') or '').strip()

    if term not in (1, 2, 3):
        return JsonResponse({'error': 'Invalid term.'}, status=400)

    # Only this adviser's own students - never another section's.
    if not Student.objects.filter(lrn=lrn, adviser_id=teacher.teacher_id).exists():
        return JsonResponse({'error': 'Student not found in your section.'}, status=404)

    StudentTermRemark.objects.update_or_create(
        lrn=lrn, term=term,
        defaults={'remark': remark, 'updated_by': teacher.teacher_id},
    )
    return JsonResponse({'status': 'ok'})


def _school_days_in_month(school_profile_id, year, month):
    """Every school day in the given month: Mon-Fri by default, adjusted by
    SchoolCalendarException rows for this school (a Saturday added as a
    make-up class, a weekday excluded as a holiday)."""
    exceptions = {
        e.date: e.is_school_day
        for e in SchoolCalendarException.objects.filter(
            school_profile_id=school_profile_id, date__year=year, date__month=month
        )
    }
    num_days = calendar.monthrange(year, month)[1]
    days = []
    for day in range(1, num_days + 1):
        d = date_cls(year, month, day)
        is_school_day = exceptions.get(d, d.weekday() < 5)
        if is_school_day:
            days.append(d)
    return days


def _holidays_payload(school_profile_id, year, month):
    """Holiday exceptions (is_school_day=False) marked for this school in the
    given month, JSON-shaped for both the initial page render and the
    add/remove-holiday AJAX responses."""
    holidays = SchoolCalendarException.objects.filter(
        school_profile_id=school_profile_id, date__year=year, date__month=month, is_school_day=False,
    ).order_by('date')
    return [
        {
            'exception_id': h.exception_id,
            'date': h.date.isoformat(),
            'label': format_date(h.date, 'M j, Y (D)'),
        }
        for h in holidays
    ]


def _grid_json_payload(teacher, section, year, month):
    """Full attendance-grid state for one month, JSON-shaped for the
    add/remove-holiday AJAX responses so the client can rebuild the day
    columns and the holiday list without a page reload."""
    school_days = _school_days_in_month(section.school_profile_id, year, month)
    students = Student.objects.filter(adviser_id=teacher.teacher_id)
    lrns = [s.lrn for s in students]
    marks = AttendanceMark.objects.filter(lrn__in=lrns, date__in=school_days)
    marks_by_key = {f'{mark.lrn}|{mark.date.isoformat()}': mark.status for mark in marks}
    return {
        'school_days': [d.isoformat() for d in school_days],
        'marks': marks_by_key,
        'holidays': _holidays_payload(section.school_profile_id, year, month),
    }


@login_required
def attendance_grid(request):
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('login')

    if teacher.role != Teacher.ROLE_ADVISER:
        messages.error(request, 'This page is only available to adviser accounts.')
        return redirect('school_dashboard')

    section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
    if not section:
        messages.error(request, 'Your profile is missing a section. Please complete your profile first.')
        return redirect('complete_profile')

    today = date_cls.today()
    try:
        year = int(request.GET.get('year', today.year))
    except ValueError:
        year = today.year
    try:
        month = int(request.GET.get('month', today.month))
    except ValueError:
        month = today.month
    if month < 1 or month > 12:
        month = today.month

    ordered_students = Student.objects.filter(adviser_id=teacher.teacher_id).order_by('surname', 'name')
    students = male_then_female(ordered_students)
    school_days = _school_days_in_month(section.school_profile_id, year, month)

    lrns = [s.lrn for s in students]
    marks = AttendanceMark.objects.filter(lrn__in=lrns, date__in=school_days)
    marks_by_key = {f'{mark.lrn}|{mark.date.isoformat()}': mark.status for mark in marks}

    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    month_start = date_cls(year, month, 1)
    month_end = date_cls(year, month, calendar.monthrange(year, month)[1])

    return render(request, 'grades/attendance_grid.html', {
        'students': students,
        'school_days': school_days,
        'marks_data': marks_by_key,
        'holidays_data': _holidays_payload(section.school_profile_id, year, month),
        'month_start': month_start.isoformat(),
        'month_end': month_end.isoformat(),
        'year': year,
        'month': month,
        'month_name': MONTH_NAMES[month - 1],
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
    })


@login_required
def attendance_grid_save(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        return JsonResponse({'error': 'Your account is not linked to a Teacher profile.'}, status=403)

    if teacher.role != Teacher.ROLE_ADVISER:
        return JsonResponse({'error': 'Adviser accounts only.'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (ValueError, UnicodeDecodeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    lrn = payload.get('lrn', '')
    date_str = payload.get('date', '')
    status = payload.get('status', '')
    remarks = payload.get('remarks') or None

    # Only this adviser's own students - never another section's.
    if not Student.objects.filter(lrn=lrn, adviser_id=teacher.teacher_id).exists():
        return JsonResponse({'error': 'Student not found in your section.'}, status=404)

    try:
        mark_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date.'}, status=400)

    if status in ('', 'present'):
        existing_marks = list(AttendanceMark.objects.filter(lrn=lrn, date=mark_date))
        for existing_mark in existing_marks:
            log_student_record_change('attendance', existing_mark.mark_id, lrn, 'deleted', teacher.teacher_id)
        if existing_marks:
            AttendanceMark.objects.filter(lrn=lrn, date=mark_date).delete()
        return JsonResponse({'status': 'present'})

    valid_statuses = {choice[0] for choice in AttendanceMark.STATUS_CHOICES}
    if status not in valid_statuses:
        return JsonResponse({'error': 'Invalid status.'}, status=400)

    mark, created = AttendanceMark.objects.update_or_create(
        lrn=lrn, date=mark_date,
        defaults={'status': status, 'remarks': remarks, 'recorded_by': teacher.teacher_id},
    )
    log_student_record_change('attendance', mark.mark_id, lrn, 'created' if created else 'updated', teacher.teacher_id)
    return JsonResponse({'status': mark.status})


def _resolve_adviser_section(request):
    """Shared Teacher/role/section resolution for the holiday endpoints,
    returning (teacher, section, None) on success or (None, None, error_response)."""
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        return None, None, JsonResponse({'error': 'Your account is not linked to a Teacher profile.'}, status=403)

    if teacher.role != Teacher.ROLE_ADVISER:
        return None, None, JsonResponse({'error': 'Adviser accounts only.'}, status=403)

    section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
    if not section:
        return None, None, JsonResponse({'error': 'Your profile is missing a section.'}, status=400)

    return teacher, section, None


@login_required
def add_holiday(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    teacher, section, error = _resolve_adviser_section(request)
    if error:
        return error

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (ValueError, UnicodeDecodeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    try:
        year = int(payload.get('year'))
        month = int(payload.get('month'))
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Invalid month.'}, status=400)

    try:
        holiday_date = datetime.strptime(payload.get('date', ''), '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date.'}, status=400)

    # Keep the added date scoped to the month currently on screen, so the
    # rebuilt grid we hand back always matches what the adviser is looking at.
    if holiday_date.year != year or holiday_date.month != month:
        return JsonResponse(
            {'error': f'Pick a date within {MONTH_NAMES[month - 1]} {year}.'}, status=400
        )

    SchoolCalendarException.objects.update_or_create(
        school_profile_id=section.school_profile_id, date=holiday_date,
        defaults={'is_school_day': False},
    )

    return JsonResponse({'ok': True, **_grid_json_payload(teacher, section, year, month)})


@login_required
def remove_holiday(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    teacher, section, error = _resolve_adviser_section(request)
    if error:
        return error

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (ValueError, UnicodeDecodeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    try:
        exception_id = int(payload.get('exception_id'))
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Invalid holiday.'}, status=400)

    # Scoped to this adviser's own school - never another school's exception
    # row, even if the id is guessed.
    exception = SchoolCalendarException.objects.filter(
        exception_id=exception_id, school_profile_id=section.school_profile_id, is_school_day=False,
    ).first()
    if not exception:
        return JsonResponse({'error': 'Holiday not found.'}, status=404)

    year, month = exception.date.year, exception.date.month
    exception.delete()

    return JsonResponse({'ok': True, **_grid_json_payload(teacher, section, year, month)})


@login_required
def view_attendance(request, lrn):
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    if lrn == 'all':
        ordered_students = Student.objects.filter(adviser_id=teacher.teacher_id).order_by('surname', 'name')
        students = male_then_female(ordered_students)

        attendance_data = {}
        for student in students:
            records = Attendance.objects.filter(lrn=student.lrn)
            month_attendance = {m: None for m in ATTENDANCE_MONTHS}
            for record in records:
                month_attendance[record.month] = record
            attendance_data[student.lrn] = month_attendance

        return render(request, 'grades/attendance_list.html', {
            'students': students,
            'attendance_data': attendance_data,
            'months': ATTENDANCE_MONTH_OPTIONS,
        })
    else:
        student = get_object_or_404(Student, lrn=lrn, adviser_id=teacher.teacher_id)
        records = Attendance.objects.filter(lrn=lrn)

        month_attendance = {m: None for m in ATTENDANCE_MONTHS}
        for record in records:
            month_attendance[record.month] = record

        return render(request, 'grades/attendance_view.html', {
            'student': student,
            'month_attendance': month_attendance,
            'months': ATTENDANCE_MONTH_OPTIONS,
        })


@login_required
def my_subject_teaching(request):
    """List every TeacherSubjectAssignment row for the requesting teacher,
    regardless of role - a Class Adviser and a Subject Teacher use this
    identically, since both can hold assignments (see school_assignments /
    adviser_subject_assignments in the school app)."""
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    assignments = TeacherSubjectAssignment.objects.filter(
        teacher_id=teacher.teacher_id
    ).order_by('assignment_id')

    sections_by_id = {
        s.section_id: s for s in Section.objects.filter(
            section_id__in=[a.section_id for a in assignments]
        )
    }
    mappings_by_id = {
        m.mapping_id: m for m in SubjectMapping.objects.filter(
            mapping_id__in=[a.mapping_id for a in assignments]
        )
    }

    assignment_rows = [
        {
            'assignment': a,
            'section': sections_by_id.get(a.section_id),
            'mapping': mappings_by_id.get(a.mapping_id),
        }
        for a in assignments
    ]

    # Ancillary tasks (Class Adviser, IT Coordinator, etc.) are self-added
    # only by Advisers/Subject Teachers - the two roles this page already
    # serves - not a general Teacher feature.
    can_manage_tasks = teacher.role in (Teacher.ROLE_ADVISER, Teacher.ROLE_SUBJECT_TEACHER)
    tasks = list(AncillaryTask.objects.filter(teacher_id=teacher.teacher_id)) if can_manage_tasks else []

    return render(request, 'grades/my_subject_teaching.html', {
        'assignment_rows': assignment_rows,
        'tasks': tasks,
        'can_manage_tasks': can_manage_tasks,
    })


@login_required
def add_ancillary_task(request):
    if request.method != 'POST':
        return redirect('my_subject_teaching')

    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    if teacher.role not in (Teacher.ROLE_ADVISER, Teacher.ROLE_SUBJECT_TEACHER):
        messages.error(request, 'Only Advisers and Subject Teachers can add ancillary tasks.')
        return redirect('my_subject_teaching')

    task_name = request.POST.get('task_name', '').strip()
    if not task_name:
        messages.error(request, 'Task name is required.')
        return redirect('my_subject_teaching')

    AncillaryTask.objects.create(teacher_id=teacher.teacher_id, task_name=task_name)
    messages.success(request, f'"{task_name}" added.')
    return redirect('my_subject_teaching')


def _get_own_ancillary_task(request, task_id):
    """Resolve the requesting Teacher and confirm the task belongs to
    them - a guessed task_id for someone else's task simply won't match.
    Returns (teacher, task, error_redirect) - error_redirect is None on
    success."""
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return None, None, redirect('dashboard')

    if teacher.role not in (Teacher.ROLE_ADVISER, Teacher.ROLE_SUBJECT_TEACHER):
        messages.error(request, 'Only Advisers and Subject Teachers can manage ancillary tasks.')
        return None, None, redirect('my_subject_teaching')

    task = AncillaryTask.objects.filter(task_id=task_id, teacher_id=teacher.teacher_id).first()
    if not task:
        messages.error(request, 'Task not found.')
        return None, None, redirect('my_subject_teaching')

    return teacher, task, None


@login_required
def edit_ancillary_task(request, task_id):
    if request.method != 'POST':
        return redirect('my_subject_teaching')

    teacher, task, error = _get_own_ancillary_task(request, task_id)
    if error:
        return error

    task_name = request.POST.get('task_name', '').strip()
    if not task_name:
        messages.error(request, 'Task name is required.')
        return redirect('my_subject_teaching')

    task.task_name = task_name
    task.save()
    messages.success(request, 'Task updated.')
    return redirect('my_subject_teaching')


@login_required
def delete_ancillary_task(request, task_id):
    if request.method != 'POST':
        return redirect('my_subject_teaching')

    teacher, task, error = _get_own_ancillary_task(request, task_id)
    if error:
        return error

    AncillaryTaskSchedule.objects.filter(task_id=task.task_id).delete()
    task_name = task.task_name
    task.delete()
    messages.success(request, f'"{task_name}" deleted.')
    return redirect('my_subject_teaching')


@login_required
def ancillary_task_schedule(request, task_id, term):
    if term not in (1, 2, 3):
        messages.error(request, 'Invalid term.')
        return redirect('my_subject_teaching')

    teacher, task, error = _get_own_ancillary_task(request, task_id)
    if error:
        return error

    if request.method == 'POST':
        task_schedule, _ = AncillaryTaskSchedule.objects.get_or_create(
            task_id=task.task_id, term=term
        )
        selected_days = set(request.POST.getlist('days'))
        task_schedule.schedule = {
            day: {
                'start': request.POST.get(f'time_start_{day}', '').strip(),
                'end': request.POST.get(f'time_end_{day}', '').strip(),
            }
            for day in WEEKDAYS if day in selected_days
        } or None
        task_schedule.save()
        messages.success(request, f'Schedule saved for "{task.task_name}".')
        return redirect('my_subject_teaching')

    task_schedule = AncillaryTaskSchedule.objects.filter(task_id=task.task_id, term=term).first()

    return render(request, 'grades/ancillary_task_schedule.html', {
        'task': task,
        'term': term,
        'schedule_dict': task_schedule.schedule_dict if task_schedule else {},
        'day_choices': [(day, WEEKDAY_LABELS[day]) for day in WEEKDAYS],
    })


WEEKDAYS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
WEEKDAY_LABELS = {'Mon': 'Monday', 'Tue': 'Tuesday', 'Wed': 'Wednesday', 'Thu': 'Thursday', 'Fri': 'Friday'}


def _score_stats(values, max_possible):
    if not values:
        return {'mean': None, 'mode': None, 'median': None, 'sd': None, 'mps': None}

    mean = statistics.mean(values)
    return {
        'mean': round(mean, 2),
        'mode': statistics.mode(values),
        'median': statistics.median(values),
        'sd': round(statistics.stdev(values), 2) if len(values) > 1 else None,
        'mps': round(mean / max_possible * 100, 2) if max_possible else None,
    }


@login_required
def subject_teaching_grades(request, assignment_id, term):
    if term not in (1, 2, 3):
        messages.error(request, 'Invalid term.')
        return redirect('my_subject_teaching')

    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return redirect('dashboard')

    # The assignment row itself must belong to this teacher - a guessed
    # assignment_id for someone else's assignment simply won't match.
    assignment = TeacherSubjectAssignment.objects.filter(
        assignment_id=assignment_id, teacher_id=teacher.teacher_id
    ).first()
    if not assignment:
        messages.error(request, 'Assignment not found.')
        return redirect('my_subject_teaching')

    section = Section.objects.filter(section_id=assignment.section_id).first()
    mapping = SubjectMapping.objects.filter(mapping_id=assignment.mapping_id).first()
    if not section or not mapping:
        messages.error(request, 'This assignment points at a missing section or subject.')
        return redirect('my_subject_teaching')

    if request.method == 'POST':
        max_score, _ = SubjectTestMaxScore.objects.get_or_create(
            assignment_id=assignment.assignment_id, term=term
        )

        selected_days = set(request.POST.getlist('days'))
        max_score.schedule = {
            day: {
                'start': request.POST.get(f'time_start_{day}', '').strip(),
                'end': request.POST.get(f'time_end_{day}', '').strip(),
            }
            for day in WEEKDAYS if day in selected_days
        } or None

        pretest_max_value = request.POST.get('pretest_max', '').strip()
        final_exam_max_value = request.POST.get('final_exam_max', '').strip()
        try:
            max_score.pretest_max = int(pretest_max_value) if pretest_max_value else None
        except ValueError:
            messages.warning(request, f'Invalid highest possible Pre-Test Score: {pretest_max_value}')
        try:
            max_score.final_exam_max = int(final_exam_max_value) if final_exam_max_value else None
        except ValueError:
            messages.warning(request, f'Invalid highest possible Final Exam Score: {final_exam_max_value}')

        for n in (1, 2, 3):
            setattr(max_score, f'pretest_llc_{n}', request.POST.get(f'pretest_llc_{n}', '').strip() or None)
            setattr(max_score, f'final_exam_llc_{n}', request.POST.get(f'final_exam_llc_{n}', '').strip() or None)

        max_score.save()

        row_count = int(request.POST.get('row_count', 0))
        saved_count = 0

        for i in range(row_count):
            lrn = request.POST.get(f'lrn_{i}', '').strip()
            if not lrn:
                continue

            # Only students actually in this assignment's section - even a
            # tampered lrn_i field can't reach a student outside it.
            if not Student.objects.filter(lrn=lrn, section_id=section.section_id).exists():
                continue

            fields = {}

            grade_value = request.POST.get(f'grade_{i}', '').strip()
            if grade_value:
                try:
                    grade = int(grade_value)
                    if grade < 60 or grade > 100:
                        messages.warning(request, f'Final Rating {grade} for {lrn} is outside 60-100 range.')
                    else:
                        fields['grade'] = grade
                except ValueError:
                    messages.warning(request, f'Invalid Final Rating value for {lrn}: {grade_value}')

            pretest_value = request.POST.get(f'pretest_{i}', '').strip()
            if pretest_value:
                try:
                    fields['pretest_score'] = int(pretest_value)
                except ValueError:
                    messages.warning(request, f'Invalid Pre-Test Score value for {lrn}: {pretest_value}')

            final_exam_value = request.POST.get(f'final_exam_{i}', '').strip()
            if final_exam_value:
                try:
                    fields['final_exam_score'] = int(final_exam_value)
                except ValueError:
                    messages.warning(request, f'Invalid Final Exam Score value for {lrn}: {final_exam_value}')

            existing = Grade.objects.filter(
                lrn=lrn, mapping_id=mapping.mapping_id, term=term
            ).first()

            # Only touch is_incomplete when it's actually changing - lets a
            # row with no other edits still save a checked/unchecked INC
            # box, without forcing every untouched row into fields.
            is_incomplete_checked = request.POST.get(f'incomplete_{i}') == 'on'
            was_incomplete = existing.is_incomplete if existing else False
            if is_incomplete_checked != was_incomplete:
                fields['is_incomplete'] = is_incomplete_checked

            if not fields:
                continue

            if existing:
                for field_name, value in fields.items():
                    setattr(existing, field_name, value)
                existing.save()
                log_student_record_change('grade', existing.grade_id, lrn, 'updated', teacher.teacher_id)
            else:
                new_grade = Grade.objects.create(
                    lrn=lrn, mapping_id=mapping.mapping_id, term=term,
                    grade=fields.get('grade'),
                    pretest_score=fields.get('pretest_score'),
                    final_exam_score=fields.get('final_exam_score'),
                    is_incomplete=fields.get('is_incomplete', False),
                    uploaded_by=teacher.teacher_id,
                )
                log_student_record_change('grade', new_grade.grade_id, lrn, 'created', teacher.teacher_id)
            saved_count += 1

        if saved_count:
            messages.success(request, f'✅ {saved_count} grades saved for {mapping.subject_name} (Term {term}).')
        return redirect('subject_teaching_grades', assignment_id=assignment.assignment_id, term=term)

    students = Student.objects.filter(section_id=section.section_id).order_by('surname', 'name')
    students = male_then_female(students)

    grades_by_lrn = {
        g.lrn: g
        for g in Grade.objects.filter(
            lrn__in=[s.lrn for s in students], mapping_id=mapping.mapping_id, term=term
        )
    }

    max_score = SubjectTestMaxScore.objects.filter(
        assignment_id=assignment.assignment_id, term=term
    ).first()

    llc_rows = [
        {
            'n': n,
            'pretest': getattr(max_score, f'pretest_llc_{n}', None),
            'final_exam': getattr(max_score, f'final_exam_llc_{n}', None),
        }
        for n in (1, 2, 3)
    ]

    pretest_values = [g.pretest_score for g in grades_by_lrn.values() if g.pretest_score is not None]
    final_exam_values = [g.final_exam_score for g in grades_by_lrn.values() if g.final_exam_score is not None]
    pretest_stats = _score_stats(pretest_values, max_score.pretest_max if max_score else None)
    final_exam_stats = _score_stats(final_exam_values, max_score.final_exam_max if max_score else None)
    schedule_dict = max_score.schedule_dict if max_score else {}

    return render(request, 'grades/subject_teaching_grades.html', {
        'assignment': assignment,
        'section': section,
        'mapping': mapping,
        'term': term,
        'students': students,
        'grades_by_lrn': grades_by_lrn,
        'max_score': max_score,
        'schedule_dict': schedule_dict,
        'llc_rows': llc_rows,
        'pretest_stats': pretest_stats,
        'final_exam_stats': final_exam_stats,
        'day_choices': [(day, WEEKDAY_LABELS[day]) for day in WEEKDAYS],
    })