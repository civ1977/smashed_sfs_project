import json

import openpyxl
from openpyxl.styles import Font, PatternFill

from collections import defaultdict

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Q, Count, Case, When, Value, IntegerField
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.db import IntegrityError, transaction
from django.utils import timezone

from datetime import date, datetime

from accounts.forms import SectionForm, SchoolProfileForm, SchoolOfficialsForm
from accounts.models import Teacher, TeacherTimeRecord, DTRCalendarException
from accounts.views import (
    build_dtr_days, DTR_TIME_FIELDS, _strip_am_pm, _dtr_calendar_exceptions, _dtr_names_match,
)
from students.models import SchoolProfile, Section, Student
from grades.models import Grade, SubjectMapping, TeacherSubjectAssignment, SchoolCalendarException, SubjectTestMaxScore, AncillaryTask, AncillaryTaskSchedule
from grades.views import MONTH_NAMES, _score_stats, WEEKDAYS
from portal.models import StudentAccount
from .models import (
    TeacherAccountAuditLog, SectionAuditLog, TimeSlot, ScheduleRequirement, ScheduleEntry,
    EnrollmentAssignment, EnrollmentApplication,
)
from . import scheduler


def _get_school_admin_teacher(request):
    """Resolve the requesting Teacher and confirm they're a registrar,
    principal, or Non-Teaching Officer (is_officer=True - a Non-Teaching
    account without that flag gets Tools/DTR only, not this).
    Returns (teacher, error_redirect) - error_redirect is None on success."""
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return None, redirect('login')

    is_officer = teacher.role == Teacher.ROLE_NON_TEACHING and teacher.is_officer
    if teacher.role not in (Teacher.ROLE_REGISTRAR, Teacher.ROLE_PRINCIPAL) and not is_officer:
        messages.error(request, 'This page is only available to registrar/principal/officer accounts.')
        return None, redirect('dashboard')

    return teacher, None


def _section_label(section):
    return f"Grade {section.grade_level}-{section.strand}-{section.section_name}"


def _mapping_matches_section(mapping, section):
    """A subject only makes sense against a section studying the same
    grade level (and, when the subject specifies one, the same strand) -
    otherwise a Grade 11 STEM subject could end up assigned against a
    Grade 12 ABM section. Shared safety-net check for every subject
    assignment entry point (the dropdowns are already narrowed to match,
    but a stale form submit could still send a mismatched mapping_id)."""
    strand_mismatch = section.strand and mapping.strand and mapping.strand != section.strand
    return mapping.grade_level == section.grade_level and not strand_mismatch


def _reassign_section_adviser(school_profile_id, section, new_adviser):
    """Point `section`'s adviser_id at `new_adviser` (a Teacher instance or
    None), clearing whatever other section that teacher previously advised
    so no teacher ends up tied to two sections at once."""
    if new_adviser is not None:
        Section.objects.filter(
            school_profile_id=school_profile_id, adviser_id=new_adviser.teacher_id
        ).exclude(section_id=section.section_id).update(adviser_id=None)
        section.adviser_id = new_adviser.teacher_id
    else:
        section.adviser_id = None
    section.save()


@login_required
def school_dashboard(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    school_profile = None
    if teacher.school_profile_id:
        school_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()

    return render(request, 'school/dashboard.html', {
        'teacher': teacher,
        'school_profile': school_profile,
    })


@login_required
def school_subject_statistics(request):
    """School-wide version of reports.subject_statistics_report: instead of
    one adviser's own section, this consolidates every section per grade
    level - Mean/Mode/Median/SD/MPS for Final Rating/Pre-Test/Final Exam,
    per subject, per term, aggregated across all sections in that grade."""
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    if not teacher.school_profile_id:
        messages.warning(request, 'Your account has no school assigned yet.')
        return render(request, 'school/subject_statistics.html', {'grade_level_rows': []})

    grade_levels = sorted(set(
        Section.objects.filter(school_profile_id=teacher.school_profile_id).values_list('grade_level', flat=True)
    ))

    grade_level_rows = []
    for grade_level in grade_levels:
        section_ids = list(Section.objects.filter(
            school_profile_id=teacher.school_profile_id, grade_level=grade_level
        ).values_list('section_id', flat=True))
        section_by_lrn = {
            s['lrn']: s['section_id']
            for s in Student.objects.filter(section_id__in=section_ids).values('lrn', 'section_id')
        }
        student_lrns = list(section_by_lrn.keys())

        subjects = SubjectMapping.objects.filter(
            school_profile_id=teacher.school_profile_id, grade_level=grade_level
        ).order_by('subject_name')

        # Multiple sections can teach the same subject under different
        # TeacherSubjectAssignment rows (one per section) - gather them all
        # per subject so a Pre-Test/Final Exam max score can still be found
        # once scores from every section are combined below.
        assignments_by_mapping = {}
        for a in TeacherSubjectAssignment.objects.filter(section_id__in=section_ids):
            assignments_by_mapping.setdefault(a.mapping_id, []).append(a)

        assignment_ids = [a.assignment_id for assignments in assignments_by_mapping.values() for a in assignments]
        max_scores = {
            (m.assignment_id, m.term): m
            for m in SubjectTestMaxScore.objects.filter(assignment_id__in=assignment_ids)
        }

        subject_scores = []
        for subject in subjects:
            grades = Grade.objects.filter(lrn__in=student_lrns, mapping_id=subject.mapping_id)
            final_by_term = {1: [], 2: [], 3: []}
            pretest_by_term = {1: [], 2: [], 3: []}
            final_exam_by_term = {1: [], 2: [], 3: []}
            # Which sections actually have a Final Rating entered for this
            # subject/term - a count of assigned sections would include ones
            # with no grades in yet, which "classes involved" shouldn't.
            final_sections_by_term = {1: set(), 2: set(), 3: set()}
            for g in grades:
                if g.grade is not None:
                    final_by_term[g.term].append(g.grade)
                    section_id = section_by_lrn.get(g.lrn)
                    if section_id is not None:
                        final_sections_by_term[g.term].add(section_id)
                if g.pretest_score is not None:
                    pretest_by_term[g.term].append(g.pretest_score)
                if g.final_exam_score is not None:
                    final_exam_by_term[g.term].append(g.final_exam_score)

            subject_scores.append({
                'subject': subject,
                'assignments': assignments_by_mapping.get(subject.mapping_id, []),
                'final_by_term': final_by_term,
                'pretest_by_term': pretest_by_term,
                'final_exam_by_term': final_exam_by_term,
                'final_sections_by_term': final_sections_by_term,
            })

        terms = []
        for term in (1, 2, 3):
            rows = []
            for entry in subject_scores:
                # Different sections' assignments can have different max
                # scores set - once their Pre-Test/Final Exam numbers are
                # combined there's no single "correct" denominator, so this
                # just uses the first one found for the term as a stand-in.
                max_score = None
                for a in entry['assignments']:
                    max_score = max_scores.get((a.assignment_id, term))
                    if max_score:
                        break
                rows.append({
                    'subject': entry['subject'],
                    'class_count': len(entry['final_sections_by_term'][term]),
                    'student_count': len(entry['final_by_term'][term]),
                    'final_rating': _score_stats(entry['final_by_term'][term], 100),
                    'pretest': _score_stats(entry['pretest_by_term'][term], max_score.pretest_max if max_score else None),
                    'final_exam': _score_stats(entry['final_exam_by_term'][term], max_score.final_exam_max if max_score else None),
                })
            terms.append({'term': term, 'rows': rows})

        grade_level_rows.append({'grade_level': grade_level, 'terms': terms})

    return render(request, 'school/subject_statistics.html', {
        'grade_level_rows': grade_level_rows,
    })


@login_required
def school_dtr_upload(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    today = date.today()

    # Any employee's DTR could have been prepared by any registrar/principal
    # at this school, not just the one viewing this page - scope by the
    # whole school's Teacher ids so "already uploaded" reflects the school,
    # not just this account's own uploads.
    school_teacher_ids = Teacher.objects.filter(
        school_profile_id=teacher.school_profile_id
    ).values_list('teacher_id', flat=True)

    uploaded_months = (
        TeacherTimeRecord.objects.filter(teacher_id__in=school_teacher_ids)
        .annotate(month_start=TruncMonth('date'))
        .values('month_start')
        .annotate(record_count=Count('record_id'))
        .order_by('-month_start')
    )
    uploaded_month_rows = [
        {
            'label': f"{MONTH_NAMES[row['month_start'].month - 1]} {row['month_start'].year}",
            'record_count': row['record_count'],
            'year': row['month_start'].year,
            'month': row['month_start'].month,
        }
        for row in uploaded_months
    ]

    calendar_exceptions = DTRCalendarException.objects.filter(
        school_profile_id=teacher.school_profile_id
    ).order_by('date')

    return render(request, 'school/dtr_upload.html', {
        'dtr_current_year': today.year,
        'dtr_current_month': today.month,
        'dtr_months': list(enumerate(MONTH_NAMES, start=1)),
        'uploaded_month_rows': uploaded_month_rows,
        'calendar_exceptions': calendar_exceptions,
        'dtr_reason_choices': DTRCalendarException.REASON_CHOICES,
        'dtr_session_choices': DTRCalendarException.SESSION_CHOICES,
    })


@login_required
def add_dtr_calendar_exception(request):
    """Marks a date Holiday or Class Suspension on the DTR calendar only -
    see DTRCalendarException for why this is kept separate from SF2's own
    SchoolCalendarException."""
    if request.method != 'POST':
        return redirect('school_dtr_upload')

    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    date_str = request.POST.get('date', '').strip()
    reason = request.POST.get('reason', '').strip()
    valid_reasons = {value for value, _ in DTRCalendarException.REASON_CHOICES}
    if reason not in valid_reasons:
        messages.error(request, 'Select a valid reason.')
        return redirect('school_dtr_upload')

    # A Holiday is always a whole day - only a Class Suspension can be
    # scoped to just one session (a suspension announced mid-morning is
    # common; a "half holiday" isn't a real DepEd concept).
    session = request.POST.get('session', '').strip()
    valid_sessions = {value for value, _ in DTRCalendarException.SESSION_CHOICES}
    if reason != DTRCalendarException.REASON_SUSPENSION or session not in valid_sessions:
        session = DTRCalendarException.SESSION_WHOLE_DAY

    try:
        exception_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Enter a valid date.')
        return redirect('school_dtr_upload')

    exception, created = DTRCalendarException.objects.update_or_create(
        school_profile_id=teacher.school_profile_id,
        date=exception_date,
        defaults={'reason': reason, 'session': session, 'created_by': teacher.teacher_id},
    )
    label = exception.get_reason_display()
    if session != DTRCalendarException.SESSION_WHOLE_DAY:
        label = f'{label} ({exception.get_session_display()})'
    verb = 'Added' if created else 'Updated'
    messages.success(request, f'{verb} {label} for {exception_date.strftime("%B %d, %Y")}.')
    return redirect('school_dtr_upload')


@login_required
def remove_dtr_calendar_exception(request, exception_id):
    if request.method != 'POST':
        return redirect('school_dtr_upload')

    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    exception = DTRCalendarException.objects.filter(
        exception_id=exception_id, school_profile_id=teacher.school_profile_id
    ).first()
    if exception:
        label = exception.get_reason_display()
        exception_date = exception.date
        exception.delete()
        messages.success(request, f'Removed {label} for {exception_date.strftime("%B %d, %Y")}.')
    else:
        messages.error(request, 'That calendar entry was not found.')

    return redirect('school_dtr_upload')


@login_required
def download_dtr_upload_template(request):
    """Blank starter workbook matching the column shape upload_dtr's own
    parser (accounts/views.py's _dtr_parse_upload_bulk) prefers - Name +
    separate Arrival/Departure date+time columns, the actual shape of this
    school's DTR app export."""
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    headers = ['Name', 'Arrival Date', 'Arrival Time', 'Departure Date', 'Departure Time', 'Total']
    sample_row = ['Dela Cruz, Juan', '2026-07-01', '08:00 AM', '2026-07-01', '05:00 PM', '9:00']

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'DTR Upload'

    sheet.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    sheet.append(sample_row)

    widths = [24, 14, 12, 14, 12, 10]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="dtr_upload_template.xlsx"'
    wb.save(response)
    return response


def _dtr_employee_status_rows(school_profile_id, year, month):
    """For every employee name found in this school's uploaded/edited DTR
    data for `year`/`month`, resolves which Teacher account it belongs to
    (same fuzzy name match as accounts.views.daily_time_record) and reports
    whether that employee has personally edited their own copy (a
    TeacherTimeRecord stamped with their own teacher_id - see
    accounts.views.save_dtr_cell/bulk_fill_dtr_lunch, as opposed to only
    appearing via a registrar's raw upload_dtr import) and whether they've
    affixed a signature. Shared by the Status page and download_dtr_pdf so
    both agree on exactly who counts as "done"."""
    school_teachers = list(Teacher.objects.filter(school_profile_id=school_profile_id))
    school_teacher_ids = [t.teacher_id for t in school_teachers]

    records = list(TeacherTimeRecord.objects.filter(
        teacher_id__in=school_teacher_ids, date__year=year, date__month=month
    ))
    employee_names = sorted(
        {r.employee_name for r in records if r.employee_name}, key=str.lower
    )

    # The same person can appear under more than one spelling in one
    # month's raw data - e.g. "Dela Cruz, Juan" from a registrar's
    # Lastname-First bulk upload alongside "Juan Dela Cruz" from the
    # teacher's own Firstname-Last account name once they self-edit.
    # self_edited/signed/status come out identical for every spelling of
    # the same matched teacher (both check matched.teacher_id /
    # matched.signature_image, not the raw string), so collapse them into
    # one row per teacher instead of double-counting the same person -
    # download_dtr_pdf already relies on this same fuzzy match to collate
    # one page per teacher; this keeps the Status page's counts and
    # checklist consistent with that.
    rows_by_teacher = {}
    unmatched_rows = []
    for name in employee_names:
        matched = next((t for t in school_teachers if _dtr_names_match(t.full_name, name)), None)
        self_edited = matched is not None and any(
            r.teacher_id == matched.teacher_id and _dtr_names_match(matched.full_name, r.employee_name)
            for r in records
        )
        signed = bool(matched and matched.signature_image)

        if matched is None:
            status = 'no_account'
        elif signed and self_edited:
            status = 'ready'
        elif self_edited:
            status = 'edited_unsigned'
        else:
            status = 'not_started'

        if matched is None:
            unmatched_rows.append({
                'employee_name': name,
                'teacher': None,
                'self_edited': False,
                'signed': False,
                'status': status,
            })
            continue

        existing = rows_by_teacher.get(matched.teacher_id)
        if existing:
            if name not in existing['employee_name'].split(' / '):
                existing['employee_name'] += f' / {name}'
            continue

        rows_by_teacher[matched.teacher_id] = {
            'employee_name': name,
            'teacher': matched,
            'self_edited': self_edited,
            'signed': signed,
            'status': status,
        }

    rows = sorted(list(rows_by_teacher.values()) + unmatched_rows, key=lambda r: r['employee_name'].lower())
    return rows, records


@login_required
def download_dtr_pdf(request, year, month):
    """Print-preview page for whichever employees the registrar checked off
    on the Status page. Renders each selected employee with the exact same
    partials/dtr_form_content.html the employee's own DTR page uses (see
    partials/dtr_form_style.html) instead of a second, separately
    maintained PDF template - the old xhtml2pdf-based version had already
    drifted from the on-screen page in real, visible ways (e.g.
    accounts.views._strip_am_pm's AM/PM-suffix trimming only ever applied
    there, never here). Producing an actual PDF file is left to the
    browser's own Print -> Save as PDF, same as the individual page's own
    Print button."""
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    status_rows, records = _dtr_employee_status_rows(teacher.school_profile_id, year, month)
    if not records:
        messages.error(request, f'No DTR data found for {MONTH_NAMES[month - 1]} {year}.')
        return redirect('school_dtr_upload')

    if request.method != 'POST':
        return redirect('dtr_status', year, month)

    selected_ids = {int(v) for v in request.POST.getlist('teacher_ids') if v.isdigit()}
    if not selected_ids:
        messages.error(request, 'Select at least one employee to download.')
        return redirect('dtr_status', year, month)

    exceptions = {
        e.date: e.is_school_day
        for e in SchoolCalendarException.objects.filter(
            school_profile_id=teacher.school_profile_id, date__year=year, date__month=month
        )
    }
    dtr_exceptions = _dtr_calendar_exceptions(teacher.school_profile_id, year, month)

    # Only employees who signed their own copy AND actually edited it
    # themselves (rather than only appearing via a registrar's raw
    # biometric import) are ever selectable on the Status page - re-checked
    # here too (status == 'ready') so a POST tampered to include someone
    # else's id can never pull them into the bulk copy. status_rows is
    # already deduped one row per matched teacher (see
    # _dtr_employee_status_rows), so this also can't double-list anyone who
    # shows up under more than one name spelling.
    forms = []
    for row in status_rows:
        et = row['teacher']
        if row['status'] != 'ready' or not et or et.teacher_id not in selected_ids:
            continue

        records_by_date = {}
        for r in records:
            if not _dtr_names_match(et.full_name, r.employee_name):
                continue
            existing = records_by_date.get(r.date)
            if existing is None or (existing.teacher_id != et.teacher_id and r.teacher_id == et.teacher_id):
                records_by_date[r.date] = r

        days = build_dtr_days(year, month, exceptions, records_by_date, dtr_exceptions)
        for entry in days:
            for field in DTR_TIME_FIELDS:
                entry[field] = _strip_am_pm(entry[field])

        forms.append({'employee_name': et.full_name, 'teacher': et, 'days': days})

    if not forms:
        messages.error(request, 'None of the selected employees are Signed & Ready.')
        return redirect('dtr_status', year, month)

    return render(request, 'school/dtr_bulk_print.html', {
        'forms': forms,
        'month_name': MONTH_NAMES[month - 1],
        'year': year,
        'month': month,
    })


@login_required
def view_employee_dtr(request, teacher_id, year, month):
    """Read-only preview of one employee's DTR for a registrar/principal to
    check before including them in a bulk selection - reuses the same
    partials/dtr_form_content.html (with readonly=True, so nothing here is
    editable) as the employee's own page and the bulk print view, so what's
    previewed here is guaranteed to match what that employee actually
    sees/signs."""
    admin_teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    target = Teacher.objects.filter(teacher_id=teacher_id, school_profile_id=admin_teacher.school_profile_id).first()
    if not target:
        messages.error(request, 'Employee not found.')
        return redirect('dtr_status', year, month)

    school_teacher_ids = Teacher.objects.filter(
        school_profile_id=admin_teacher.school_profile_id
    ).values_list('teacher_id', flat=True)

    exceptions = {
        e.date: e.is_school_day
        for e in SchoolCalendarException.objects.filter(
            school_profile_id=admin_teacher.school_profile_id, date__year=year, date__month=month
        )
    }
    dtr_exceptions = _dtr_calendar_exceptions(admin_teacher.school_profile_id, year, month)

    records_by_date = {}
    for r in TeacherTimeRecord.objects.filter(teacher_id__in=school_teacher_ids, date__year=year, date__month=month):
        if not _dtr_names_match(target.full_name, r.employee_name):
            continue
        existing = records_by_date.get(r.date)
        if existing is None or (existing.teacher_id != target.teacher_id and r.teacher_id == target.teacher_id):
            records_by_date[r.date] = r

    days = build_dtr_days(year, month, exceptions, records_by_date, dtr_exceptions)
    for entry in days:
        for field in DTR_TIME_FIELDS:
            entry[field] = _strip_am_pm(entry[field])

    return render(request, 'school/dtr_view.html', {
        'teacher': target,
        'days': days,
        'employee_name': target.full_name,
        'month_name': MONTH_NAMES[month - 1],
        'year': year,
        'month': month,
    })


@login_required
def dtr_status(request, year, month):
    """Per-employee status for one uploaded month: every name found in that
    month's uploaded/edited DTR data (the "1st column" of the original
    excel/csv, i.e. TeacherTimeRecord.employee_name), and whether they've
    personally edited their own copy and affixed a signature - the exact
    same criteria download_dtr_pdf uses to decide who gets collated into
    the bulk PDF, so this page doubles as a checklist of who's ready."""
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    status_rows, records = _dtr_employee_status_rows(teacher.school_profile_id, year, month)
    if not records:
        messages.error(request, f'No DTR data found for {MONTH_NAMES[month - 1]} {year}.')
        return redirect('school_dtr_upload')

    ready_count = sum(1 for r in status_rows if r['status'] == 'ready')

    return render(request, 'school/dtr_status.html', {
        'status_rows': status_rows,
        'ready_count': ready_count,
        'month_name': MONTH_NAMES[month - 1],
        'year': year,
        'month': month,
    })


@login_required
def delete_dtr_month(request, year, month):
    if request.method != 'POST':
        return redirect('school_dtr_upload')

    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    school_teacher_ids = Teacher.objects.filter(
        school_profile_id=teacher.school_profile_id
    ).values_list('teacher_id', flat=True)

    deleted_count, _ = TeacherTimeRecord.objects.filter(
        teacher_id__in=school_teacher_ids, date__year=year, date__month=month
    ).delete()

    target_label = f"{MONTH_NAMES[month - 1]} {year}"
    if deleted_count:
        messages.success(request, f'Deleted {deleted_count} DTR record(s) for {target_label}.')
    else:
        messages.error(request, f'No DTR data found for {target_label}.')
    return redirect('school_dtr_upload')


@login_required
def school_student_list(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    query = request.GET.get('q', '').strip()
    student_rows = []
    school_year_options = []
    selected_profile_id = teacher.school_profile_id

    if not teacher.school_profile_id:
        messages.warning(request, 'Your account has no school assigned yet.')
    else:
        # Every SchoolProfile sharing this school's DepEd school_id is a
        # different school_year's own copy of that school (see
        # SchoolProfile.school_id vs. .profile_id) - list them all here so a
        # registrar can search a prior year's own sections/students instead
        # of only whichever profile_id their account currently points at.
        current_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()
        if current_profile:
            school_year_options = list(
                SchoolProfile.objects.filter(school_id=current_profile.school_id).order_by('-school_year')
            )
            option_ids = {p.profile_id for p in school_year_options}
            requested_profile_id = request.GET.get('profile_id', '').strip()
            if requested_profile_id.isdigit() and int(requested_profile_id) in option_ids:
                selected_profile_id = int(requested_profile_id)

        if query:
            sections_by_id = {
                s.section_id: s for s in Section.objects.filter(school_profile_id=selected_profile_id)
            }
            students = Student.objects.filter(
                section_id__in=sections_by_id.keys()
            ).filter(
                Q(lrn__icontains=query) | Q(surname__icontains=query) | Q(name__icontains=query)
            ).order_by('surname', 'name')

            adviser_ids = {s.adviser_id for s in sections_by_id.values() if s.adviser_id}
            advisers_by_id = {t.teacher_id: t for t in Teacher.objects.filter(teacher_id__in=adviser_ids)}

            student_rows = []
            for s in students:
                section = sections_by_id.get(s.section_id)
                student_rows.append({
                    'student': s,
                    'section': section,
                    'adviser': advisers_by_id.get(section.adviser_id) if section else None,
                })

    return render(request, 'school/students.html', {
        'query': query,
        'student_rows': student_rows,
        'school_year_options': school_year_options,
        'selected_profile_id': selected_profile_id,
    })


@login_required
def school_sections(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    if not teacher.school_profile_id:
        messages.warning(request, 'Your account has no school assigned yet.')
        return render(request, 'school/sections.html', {'section_rows': []})

    sections = Section.objects.filter(
        school_profile_id=teacher.school_profile_id
    ).order_by('grade_level', 'section_name')

    adviser_ids = [s.adviser_id for s in sections if s.adviser_id]
    advisers_by_id = {t.teacher_id: t for t in Teacher.objects.filter(teacher_id__in=adviser_ids)}
    student_counts = {
        s.section_id: Student.objects.filter(section_id=s.section_id).count() for s in sections
    }

    section_rows = [
        {
            'section': s,
            'adviser': advisers_by_id.get(s.adviser_id),
            'student_count': student_counts.get(s.section_id, 0),
        }
        for s in sections
    ]

    all_teachers = Teacher.objects.filter(
        school_profile_id=teacher.school_profile_id, role=Teacher.ROLE_ADVISER
    ).order_by('full_name')

    teachers_by_id = {t.teacher_id: t for t in Teacher.objects.filter(school_profile_id=teacher.school_profile_id)}
    audit_log = SectionAuditLog.objects.filter(school_profile_id=teacher.school_profile_id)[:50]
    audit_log_rows = [
        {'log': entry, 'performed_by': teachers_by_id.get(entry.performed_by)}
        for entry in audit_log
    ]

    return render(request, 'school/sections.html', {
        'section_rows': section_rows,
        'all_teachers': all_teachers,
        'audit_log_rows': audit_log_rows,
    })


@login_required
def school_section_students(request, section_id):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    section = Section.objects.filter(
        section_id=section_id, school_profile_id=teacher.school_profile_id
    ).first()
    if not section:
        messages.error(request, 'Section not found.')
        return redirect('school_sections')

    adviser = Teacher.objects.filter(teacher_id=section.adviser_id).first() if section.adviser_id else None
    # All MALE students first, then all FEMALE - alphabetical by name within
    # each group, rather than sex just happening to sort alphabetically
    # (which would put FEMALE first).
    students = Student.objects.filter(section_id=section.section_id).order_by(
        Case(
            When(sex='MALE', then=Value(0)),
            When(sex='FEMALE', then=Value(1)),
            default=Value(2),
            output_field=IntegerField(),
        ),
        'surname', 'name',
    )

    return render(request, 'school/section_students.html', {
        'section': section,
        'adviser': adviser,
        'students': students,
    })


@login_required
def reassign_section_adviser(request, section_id):
    if request.method != 'POST':
        return redirect('school_sections')

    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    section = Section.objects.filter(
        section_id=section_id, school_profile_id=teacher.school_profile_id
    ).first()
    if not section:
        messages.error(request, 'Section not found.')
        return redirect('school_sections')

    adviser_teacher_id = request.POST.get('adviser_teacher_id', '').strip()
    if not adviser_teacher_id:
        _reassign_section_adviser(teacher.school_profile_id, section, None)
        messages.success(request, f'{section.section_name} is now unassigned.')
        SectionAuditLog.objects.create(
            school_profile_id=teacher.school_profile_id,
            section_label=_section_label(section),
            action=SectionAuditLog.ACTION_EDITED,
            detail='Adviser unassigned',
            performed_by=teacher.teacher_id,
        )
    else:
        adviser = Teacher.objects.filter(
            teacher_id=adviser_teacher_id, school_profile_id=teacher.school_profile_id
        ).first()
        if not adviser:
            messages.error(request, 'Teacher not found in this school.')
        else:
            _reassign_section_adviser(teacher.school_profile_id, section, adviser)
            messages.success(request, f'✅ {section.section_name} adviser reassigned to {adviser.full_name}.')
            SectionAuditLog.objects.create(
                school_profile_id=teacher.school_profile_id,
                section_label=_section_label(section),
                action=SectionAuditLog.ACTION_EDITED,
                detail=f'Adviser reassigned to {adviser.full_name}',
                performed_by=teacher.teacher_id,
            )

    return redirect('school_sections')


@login_required
def delete_section(request, section_id):
    if request.method != 'POST':
        return redirect('school_sections')

    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    section = Section.objects.filter(
        section_id=section_id, school_profile_id=teacher.school_profile_id
    ).first()
    if not section:
        messages.error(request, 'Section not found.')
        return redirect('school_sections')

    # A section with students or subject-teaching assignments still pointing
    # at it can't be deleted - there's no cascading delete in this schema
    # (see CLAUDE.md), so removing it would silently orphan those rows.
    student_count = Student.objects.filter(section_id=section.section_id).count()
    assignment_count = TeacherSubjectAssignment.objects.filter(section_id=section.section_id).count()
    if student_count or assignment_count:
        messages.error(
            request,
            f'Cannot delete {section.section_name}: it still has {student_count} student(s) and '
            f'{assignment_count} subject-teaching assignment(s). Move or remove those first.'
        )
        return redirect('school_sections')

    label = _section_label(section)
    section.delete()

    SectionAuditLog.objects.create(
        school_profile_id=teacher.school_profile_id,
        section_label=label,
        action=SectionAuditLog.ACTION_DELETED,
        detail='',
        performed_by=teacher.teacher_id,
    )
    messages.success(request, f'✅ {label} deleted.')
    return redirect('school_sections')


@login_required
def school_accounts(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    if not teacher.school_profile_id:
        messages.warning(request, 'Your account has no school assigned yet.')
        return render(request, 'school/accounts.html', {'account_rows': [], 'sections': []})

    teachers = Teacher.objects.filter(school_profile_id=teacher.school_profile_id).order_by('full_name')
    sections = Section.objects.filter(school_profile_id=teacher.school_profile_id).order_by('grade_level', 'section_name')
    sections_by_adviser = {s.adviser_id: s for s in sections if s.adviser_id}

    account_rows = [
        {'teacher': t, 'section': sections_by_adviser.get(t.teacher_id)}
        for t in teachers
    ]

    teachers_by_id = {t.teacher_id: t for t in teachers}
    audit_log = TeacherAccountAuditLog.objects.filter(
        target_teacher_id__in=teachers_by_id.keys()
    )[:50]
    audit_log_rows = [
        {
            'log': entry,
            'target': teachers_by_id.get(entry.target_teacher_id),
            'performed_by': teachers_by_id.get(entry.performed_by),
        }
        for entry in audit_log
    ]

    return render(request, 'school/accounts.html', {
        'account_rows': account_rows,
        'sections': sections,
        'audit_log_rows': audit_log_rows,
    })


@login_required
def toggle_teacher_active(request, teacher_id):
    if request.method != 'POST':
        return redirect('school_accounts')

    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    target = Teacher.objects.filter(
        teacher_id=teacher_id, school_profile_id=teacher.school_profile_id
    ).first()
    if not target:
        messages.error(request, 'Teacher not found in this school.')
        return redirect('school_accounts')

    target.is_active = not target.is_active
    target.save()

    # Teacher.is_active is this app's own record of standing; the linked
    # auth.User.is_active is what Django's authenticate() actually checks,
    # so keep them in sync - that's what blocks (or restores) future login.
    # This never touches Student/Grade rows, so historical data the teacher
    # already entered stays intact either way.
    if target.user_id:
        target.user.is_active = target.is_active
        target.user.save()

    status = 'activated' if target.is_active else 'deactivated'
    TeacherAccountAuditLog.objects.create(
        target_teacher_id=target.teacher_id,
        action=status,
        performed_by=teacher.teacher_id,
    )
    messages.success(request, f'✅ {target.full_name} {status}.')
    return redirect('school_accounts')


@login_required
def toggle_teacher_officer(request, teacher_id):
    """Promote/demote an existing Non-Teaching account between Officer
    (full school-admin access) and Non-Officer (Tools/DTR only) - the
    is_officer flag is otherwise only set once, at registration, so this
    is the only way to change it for an account created before that
    choice existed or that was registered with the wrong one."""
    if request.method != 'POST':
        return redirect('school_accounts')

    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    target = Teacher.objects.filter(
        teacher_id=teacher_id, school_profile_id=teacher.school_profile_id,
        role=Teacher.ROLE_NON_TEACHING,
    ).first()
    if not target:
        messages.error(request, 'Non-teaching account not found in this school.')
        return redirect('school_accounts')

    target.is_officer = not target.is_officer
    target.save()

    status = 'promoted_officer' if target.is_officer else 'demoted_officer'
    TeacherAccountAuditLog.objects.create(
        target_teacher_id=target.teacher_id,
        action=status,
        performed_by=teacher.teacher_id,
    )
    label = 'promoted to Officer' if target.is_officer else 'demoted to Non-Officer'
    messages.success(request, f'✅ {target.full_name} {label}.')
    return redirect('school_accounts')


@login_required
def reassign_teacher_section(request, teacher_id):
    if request.method != 'POST':
        return redirect('school_accounts')

    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    target = Teacher.objects.filter(
        teacher_id=teacher_id, school_profile_id=teacher.school_profile_id
    ).first()
    if not target:
        messages.error(request, 'Teacher not found in this school.')
        return redirect('school_accounts')

    section_id = request.POST.get('section_id', '').strip()
    if not section_id:
        Section.objects.filter(
            school_profile_id=teacher.school_profile_id, adviser_id=target.teacher_id
        ).update(adviser_id=None)
        messages.success(request, f'{target.full_name} unassigned from their section.')
        return redirect('school_accounts')

    section = Section.objects.filter(
        section_id=section_id, school_profile_id=teacher.school_profile_id
    ).first()
    if not section:
        messages.error(request, 'Section not found.')
        return redirect('school_accounts')

    _reassign_section_adviser(teacher.school_profile_id, section, target)
    messages.success(request, f'✅ {target.full_name} is now adviser of {section.section_name}.')
    return redirect('school_accounts')


@login_required
def school_assignments(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    if not teacher.school_profile_id:
        messages.warning(request, 'Your account has no school assigned yet.')
        return render(request, 'school/assignments.html', {
            'teacher_rows': [], 'subject_teachers': [], 'sections': [], 'mappings': [],
        })

    if request.method == 'POST':
        subject_teacher_id = request.POST.get('teacher_id', '').strip()
        section_id = request.POST.get('section_id', '').strip()
        mapping_id = request.POST.get('mapping_id', '').strip()

        if not (subject_teacher_id and section_id and mapping_id):
            messages.error(request, 'Select a subject teacher, section, and subject.')
            return redirect('school_assignments')

        subject_teacher = Teacher.objects.filter(
            teacher_id=subject_teacher_id,
            school_profile_id=teacher.school_profile_id,
            role__in=(Teacher.ROLE_SUBJECT_TEACHER, Teacher.ROLE_ADVISER),
        ).first()
        section = Section.objects.filter(
            section_id=section_id, school_profile_id=teacher.school_profile_id
        ).first()
        mapping = SubjectMapping.objects.filter(
            mapping_id=mapping_id, school_profile_id=teacher.school_profile_id
        ).first()

        if not (subject_teacher and section and mapping):
            messages.error(request, 'Selected teacher, section, or subject not found in this school.')
            return redirect('school_assignments')

        if not _mapping_matches_section(mapping, section):
            messages.error(
                request,
                f'{mapping.subject_name} (Grade {mapping.grade_level} {mapping.strand}) '
                f'doesn\'t match {section.grade_level}-{section.strand}-{section.section_name}.'
            )
            return redirect('school_assignments')

        TeacherSubjectAssignment.objects.create(
            teacher_id=subject_teacher.teacher_id,
            section_id=section.section_id,
            mapping_id=mapping.mapping_id,
        )
        messages.success(
            request,
            f'✅ {subject_teacher.full_name} assigned to {mapping.subject_name} in {section.section_name}.'
        )
        return redirect('school_assignments')

    subject_teachers = Teacher.objects.filter(
        school_profile_id=teacher.school_profile_id,
        role__in=(Teacher.ROLE_SUBJECT_TEACHER, Teacher.ROLE_ADVISER),
    ).order_by('full_name')

    sections = Section.objects.filter(
        school_profile_id=teacher.school_profile_id
    ).order_by('grade_level', 'strand', 'section_name')

    mappings = SubjectMapping.objects.filter(
        school_profile_id=teacher.school_profile_id, is_active=True
    ).order_by('grade_level', 'strand', 'subject_name')

    sections_by_id = {s.section_id: s for s in sections}
    mappings_by_id = {m.mapping_id: m for m in mappings}

    assignments = TeacherSubjectAssignment.objects.filter(
        teacher_id__in=[t.teacher_id for t in subject_teachers]
    ).order_by('assignment_id')

    assignments_by_teacher = {}
    for a in assignments:
        assignments_by_teacher.setdefault(a.teacher_id, []).append({
            'assignment': a,
            'section': sections_by_id.get(a.section_id),
            'mapping': mappings_by_id.get(a.mapping_id),
        })

    teacher_rows = [
        {'teacher': t, 'assignments': assignments_by_teacher.get(t.teacher_id, [])}
        for t in subject_teachers
    ]

    return render(request, 'school/assignments.html', {
        'teacher_rows': teacher_rows,
        'subject_teachers': subject_teachers,
        'sections': sections,
        'mappings': mappings,
    })


def _format_ancillary_schedule(schedule):
    """Same rendering as reports/views.py's _format_schedule, kept as its
    own small copy here rather than importing a module-private helper
    across apps for a single 6-line formatter."""
    if not schedule:
        return '—'
    parts = []
    for day in WEEKDAYS:
        entry = schedule.get(day)
        if not entry:
            continue
        start = entry.get('start') or '?'
        end = entry.get('end') or '?'
        parts.append(f'{day} {start}-{end}')
    return ', '.join(parts) if parts else '—'


@login_required
def school_ancillary_tasks(request):
    """School-wide, read-only view of every Adviser/Subject Teacher's
    ancillary tasks (Class Adviser, IT Coordinator, etc.) and their
    schedules, consolidated in one place for SF7 preparation - the tasks
    themselves are self-added by each teacher on their own My Subject
    Teaching page (grades/views.py's my_subject_teaching)."""
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    if not teacher.school_profile_id:
        messages.warning(request, 'Your account has no school assigned yet.')
        return render(request, 'school/ancillary_tasks.html', {'teacher_rows': []})

    staff = Teacher.objects.filter(
        school_profile_id=teacher.school_profile_id,
        role__in=(Teacher.ROLE_ADVISER, Teacher.ROLE_SUBJECT_TEACHER),
    ).order_by('full_name')

    tasks_by_teacher = {}
    for t in AncillaryTask.objects.filter(teacher_id__in=[s.teacher_id for s in staff]):
        tasks_by_teacher.setdefault(t.teacher_id, []).append(t)

    task_ids = [t.task_id for tasks in tasks_by_teacher.values() for t in tasks]
    schedules_by_task = {}
    for sched in AncillaryTaskSchedule.objects.filter(task_id__in=task_ids):
        schedules_by_task.setdefault(sched.task_id, {})[sched.term] = sched.schedule

    teacher_rows = []
    for s in staff:
        task_rows = []
        for t in tasks_by_teacher.get(s.teacher_id, []):
            term_schedules = schedules_by_task.get(t.task_id, {})
            task_rows.append({
                'task_name': t.task_name,
                'schedule': ' | '.join(
                    f'T{term}: {_format_ancillary_schedule(term_schedules.get(term))}'
                    for term in (1, 2, 3)
                ),
            })
        if task_rows:
            teacher_rows.append({'teacher': s, 'tasks': task_rows})

    return render(request, 'school/ancillary_tasks.html', {'teacher_rows': teacher_rows})


# ---------------------------------------------------------------------------
# Enrollment - real persistence: EnrollmentAssignment/EnrollmentApplication
# (school/models.py) and SchoolProfile.enrollment_enabled (students/models.py).
# An application deliberately holds the applicant's full submitted info
# directly rather than a Student FK - Student.section_id/adviser_id are
# required fields an enrollee doesn't have yet, so turning an application
# into a real Student row (and picking their section) stays a separate,
# manual step for a registrar via the existing Sections/Students pages.
# ---------------------------------------------------------------------------

def _scope_label(scope_code):
    return 'Whole School' if scope_code == 'all' else f'Grade {scope_code}'


def _assignees_for_grade(assignments, grade_level):
    """Names of everyone assigned to process this grade level, from a list
    of EnrollmentAssignment rows already annotated with `.teacher_name` (see
    _assignments_with_names) - a Whole School (scope_code 'all') assignment
    counts for every grade level. Order-preserving de-dupe so a name doesn't
    repeat if assigned both ways."""
    names = [a.teacher_name for a in assignments if a.scope_code in ('all', grade_level)]
    seen = set()
    result = []
    for name in names:
        if name not in seen:
            seen.add(name)
            result.append(name)
    return result


def _assignments_with_names(school_profile_id):
    """EnrollmentAssignment has no FK to Teacher (no-FK convention across
    this codebase), so stamp each row with .teacher_name/.scope_label here
    rather than in the template, which can't do the lookup itself."""
    assignments = list(EnrollmentAssignment.objects.filter(school_profile_id=school_profile_id))
    names_by_id = {
        t.teacher_id: t.full_name for t in Teacher.objects.filter(
            teacher_id__in=[a.teacher_id for a in assignments]
        )
    }
    for a in assignments:
        a.teacher_name = names_by_id.get(a.teacher_id, 'Unknown')
        a.scope_label = _scope_label(a.scope_code)
    return assignments


def _get_any_teacher(request):
    """Resolve the requesting Teacher without a role restriction - used by
    the enrollment-processing views, which any assigned personnel (adviser,
    subject teacher, non-teaching staff) or school admin can reach."""
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return None, redirect('dashboard')
    return teacher, None


def _send_enrollment_email(to_email, subject, message):
    """Best-effort notification - fail_silently=True so a misconfigured
    SMTP server never blocks the actual enrollment action (toggle, assign,
    schedule, decide) that triggered it."""
    if not to_email:
        return
    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [to_email], fail_silently=True)


@login_required
def school_enrollment(request):
    """Registrar/Principal/Non-Teaching Officer view: enable/disable, assign
    personnel (whole school or per grade level), and the applicant queue."""
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    school_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'toggle_enrollment' and school_profile:
            school_profile.enrollment_enabled = not school_profile.enrollment_enabled
            school_profile.save(update_fields=['enrollment_enabled'])
            messages.success(request, 'Enrollment is now {}.'.format(
                'open' if school_profile.enrollment_enabled else 'closed'
            ))
        elif action == 'assign_personnel':
            teacher_id = request.POST.get('teacher_id')
            scope_code = request.POST.get('scope')
            assigned_teacher = Teacher.objects.filter(
                teacher_id=teacher_id, school_profile_id=teacher.school_profile_id
            ).first()
            if assigned_teacher and scope_code:
                _assignment, created = EnrollmentAssignment.objects.get_or_create(
                    school_profile_id=teacher.school_profile_id,
                    teacher_id=assigned_teacher.teacher_id,
                    scope_code=scope_code,
                    defaults={'assigned_by': teacher.teacher_id},
                )
                scope_label = _scope_label(scope_code)
                if created:
                    messages.success(request, f'{assigned_teacher.full_name} added to {scope_label}.')
                else:
                    messages.info(request, f'{assigned_teacher.full_name} is already assigned to {scope_label}.')
            else:
                messages.error(request, 'Select both a personnel and a scope.')
        elif action == 'unassign_personnel':
            scope_code = request.POST.get('scope_code')
            unassign_teacher_id = request.POST.get('teacher_id')
            deleted, _ = EnrollmentAssignment.objects.filter(
                school_profile_id=teacher.school_profile_id,
                scope_code=scope_code,
                teacher_id=unassign_teacher_id,
            ).delete()
            if deleted:
                removed = Teacher.objects.filter(teacher_id=unassign_teacher_id).first()
                messages.success(
                    request,
                    f'{removed.full_name if removed else "Personnel"} removed from {_scope_label(scope_code)}.',
                )
        return redirect('school_enrollment')

    staff = Teacher.objects.filter(
        school_profile_id=teacher.school_profile_id,
        role__in=(Teacher.ROLE_NON_TEACHING, Teacher.ROLE_ADVISER, Teacher.ROLE_SUBJECT_TEACHER),
    ).order_by('full_name')
    grade_levels = sorted(set(
        Section.objects.filter(school_profile_id=teacher.school_profile_id).values_list('grade_level', flat=True)
    ))

    applications = EnrollmentApplication.objects.filter(school_profile_id=teacher.school_profile_id)
    queue_grade_levels = sorted(set(applications.values_list('grade_level', flat=True)))
    grade_queues = [
        {
            'grade_level': gl,
            'total': applications.filter(grade_level=gl).count(),
            'pending': applications.filter(grade_level=gl, status=EnrollmentApplication.STATUS_PENDING).count(),
        }
        for gl in queue_grade_levels
    ]

    return render(request, 'school/enrollment.html', {
        'enrollment_enabled': bool(school_profile and school_profile.enrollment_enabled),
        'staff': staff,
        'grade_levels': grade_levels,
        'assignments': _assignments_with_names(teacher.school_profile_id),
        'grade_queues': grade_queues,
    })


@login_required
def school_enrollment_queue(request, grade_level):
    """Applicant queue for a single grade level, reached via the per-grade
    buttons on the main Enrollment page."""
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    assignments = _assignments_with_names(teacher.school_profile_id)
    applications = EnrollmentApplication.objects.filter(
        school_profile_id=teacher.school_profile_id, grade_level=grade_level
    )
    queue = [
        {'application': app, 'assigned_display': ', '.join(_assignees_for_grade(assignments, grade_level)) or None}
        for app in applications
    ]

    return render(request, 'school/enrollment_queue.html', {
        'grade_level': grade_level,
        'queue': queue,
    })


@login_required
def school_enrollment_my_queue(request):
    """Landing page for any assigned personnel (adviser/subject teacher/
    non-teaching staff) reached from their Tools tab: applications for any
    grade level (or the whole school) currently assigned to them - a grade
    can have several assignees, so this isn't a 1:1 "owner" match."""
    teacher, error = _get_any_teacher(request)
    if error:
        return error

    my_scopes = set(EnrollmentAssignment.objects.filter(
        school_profile_id=teacher.school_profile_id, teacher_id=teacher.teacher_id
    ).values_list('scope_code', flat=True))

    applications = EnrollmentApplication.objects.filter(school_profile_id=teacher.school_profile_id)
    if 'all' not in my_scopes:
        applications = applications.filter(grade_level__in=my_scopes)

    return render(request, 'school/enrollment_my_queue.html', {
        'teacher': teacher,
        'queue': applications,
    })


@login_required
def school_enrollment_process(request, application_id):
    """Any assigned personnel (adviser/subject teacher/non-teaching staff)
    or school admin can process an application: schedule a personal
    appearance, then decide Enrolled / Reject / Other once they show up."""
    teacher, error = _get_any_teacher(request)
    if error:
        return error

    is_school_admin = bool(teacher and (
        teacher.role in (Teacher.ROLE_REGISTRAR, Teacher.ROLE_PRINCIPAL)
        or (teacher.role == Teacher.ROLE_NON_TEACHING and teacher.is_officer)
    ))

    application = EnrollmentApplication.objects.filter(
        application_id=application_id, school_profile_id=teacher.school_profile_id
    ).first()
    if not application:
        messages.error(request, 'Application not found.')
        return redirect('school_enrollment' if is_school_admin else 'school_enrollment_my_queue')

    applicant_email = User.objects.filter(
        id__in=StudentAccount.objects.filter(lrn=application.lrn).values_list('user_id', flat=True)
    ).values_list('email', flat=True).first()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'schedule':
            appearance_date = request.POST.get('appearance_date')
            appearance_time = request.POST.get('appearance_time')
            if appearance_date and appearance_time:
                application.appearance_date = datetime.strptime(appearance_date, '%Y-%m-%d').date()
                application.appearance_time = datetime.strptime(appearance_time, '%H:%M').time()
                application.scheduled_by = teacher.teacher_id
                application.scheduled_at = timezone.now()
                if application.status == EnrollmentApplication.STATUS_PENDING:
                    application.status = EnrollmentApplication.STATUS_SCHEDULED
                application.save()
                when = f"{application.appearance_date.strftime('%B %d, %Y')} at {application.appearance_time.strftime('%I:%M %p').lstrip('0')}"
                _send_enrollment_email(
                    applicant_email,
                    'Enrollment: Personal Appearance Schedule',
                    f'Hello {application.given_name},\n\n'
                    f'Please report in person for your enrollment on {when} to complete your application.\n\n'
                    'Thank you.',
                )
                messages.success(request, 'Personal appearance scheduled and the applicant has been notified by email.')
            else:
                messages.error(request, 'Select both a date and a time.')
        elif action in ('enrolled', 'rejected', 'other'):
            application.status = action
            application.decided_by = teacher.teacher_id
            application.decided_at = timezone.now()
            application.save()
            messages.success(request, f'Application marked as "{action}".')
        return redirect('school_enrollment_process', application_id=application_id)

    assignments = _assignments_with_names(teacher.school_profile_id)
    assigned_names = _assignees_for_grade(assignments, application.grade_level)

    return render(request, 'school/enrollment_process.html', {
        'application': application,
        'assigned_display': ', '.join(assigned_names) if assigned_names else None,
        'is_school_admin': is_school_admin,
    })


@login_required
def remove_teacher_subject_assignment(request, assignment_id):
    if request.method != 'POST':
        return redirect('school_assignments')

    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    assignment = TeacherSubjectAssignment.objects.filter(assignment_id=assignment_id).first()
    if not assignment:
        messages.error(request, 'Assignment not found.')
        return redirect('school_assignments')

    # Scope by the assignment's own teacher rather than a stored school id
    # (the model has none) - same manual-FK discipline as everywhere else.
    assignment_teacher = Teacher.objects.filter(
        teacher_id=assignment.teacher_id,
        school_profile_id=teacher.school_profile_id,
        role__in=(Teacher.ROLE_SUBJECT_TEACHER, Teacher.ROLE_ADVISER),
    ).first()
    if not assignment_teacher:
        messages.error(request, 'Assignment not found in this school.')
        return redirect('school_assignments')

    assignment.delete()
    messages.success(request, f'Assignment removed from {assignment_teacher.full_name}.')
    return redirect('school_assignments')


def _get_adviser_with_section(request):
    """Resolve the requesting Teacher, confirm they're a Class Adviser (not
    _get_school_admin_teacher, which explicitly excludes advisers), and
    resolve their own Section. Returns (teacher, section, error_redirect) -
    error_redirect is None on success."""
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to a Teacher profile.')
        return None, None, redirect('login')

    if teacher.role != Teacher.ROLE_ADVISER:
        messages.error(request, 'This page is only available to Class Advisers.')
        return None, None, redirect('dashboard')

    section = Section.objects.filter(adviser_id=teacher.teacher_id).first()
    if not section:
        messages.warning(request, 'Complete your section setup first.')
        return None, None, redirect('complete_profile')

    return teacher, section, None


@login_required
def edit_my_section(request):
    teacher, section, error = _get_adviser_with_section(request)
    if error:
        return error

    school_profile = SchoolProfile.objects.filter(profile_id=section.school_profile_id).first()

    if request.method == 'POST':
        form = SectionForm(request.POST, instance=section, prefix='section')
        officials_form = SchoolOfficialsForm(
            request.POST, instance=school_profile, prefix='officials'
        ) if school_profile else None

        if form.is_valid() and (officials_form is None or officials_form.is_valid()):
            form.save()
            if officials_form:
                officials_form.save()
            messages.success(request, '✅ Classroom details updated.')
            return redirect('tools')
    else:
        form = SectionForm(instance=section, prefix='section')
        officials_form = SchoolOfficialsForm(
            instance=school_profile, prefix='officials'
        ) if school_profile else None

    return render(request, 'school/edit_my_section.html', {
        'section': section,
        'section_form': form,
        'officials_form': officials_form,
    })


@login_required
def edit_school_profile(request):
    """Registrar/principal-only edit of the school-wide profile (name,
    region/division/district/municipality, registrar/principal/SDS names) -
    the same SchoolProfileForm used to create one during profile setup,
    reused here for editing an existing one."""
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    school_profile = SchoolProfile.objects.filter(profile_id=teacher.school_profile_id).first()
    if not school_profile:
        messages.error(request, 'No school profile found. Please complete your profile first.')
        return redirect('complete_profile')

    if request.method == 'POST':
        form = SchoolProfileForm(request.POST, instance=school_profile)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ School data updated.')
            return redirect('tools')
    else:
        form = SchoolProfileForm(instance=school_profile)

    return render(request, 'school/edit_school_profile.html', {
        'school_profile': school_profile,
        'school_profile_form': form,
    })


@login_required
def adviser_subject_assignments(request):
    teacher, section, error = _get_adviser_with_section(request)
    if error:
        return error

    if request.method == 'POST':
        subject_teacher_id = request.POST.get('teacher_id', '').strip()
        mapping_id = request.POST.get('mapping_id', '').strip()

        if not (subject_teacher_id and mapping_id):
            messages.error(request, 'Select a teacher and a subject.')
            return redirect('adviser_subject_assignments')

        subject_teacher = Teacher.objects.filter(
            teacher_id=subject_teacher_id,
            school_profile_id=teacher.school_profile_id,
            role__in=(Teacher.ROLE_ADVISER, Teacher.ROLE_SUBJECT_TEACHER),
        ).first()
        mapping = SubjectMapping.objects.filter(
            mapping_id=mapping_id, school_profile_id=teacher.school_profile_id
        ).first()

        if not (subject_teacher and mapping):
            messages.error(request, 'Selected teacher or subject not found in this school.')
            return redirect('adviser_subject_assignments')

        if not _mapping_matches_section(mapping, section):
            messages.error(
                request,
                f'{mapping.subject_name} (Grade {mapping.grade_level} {mapping.strand}) '
                f'doesn\'t match your section {section.grade_level}-{section.strand}-{section.section_name}.'
            )
            return redirect('adviser_subject_assignments')

        # Never trust a section_id from the request - always the adviser's own.
        TeacherSubjectAssignment.objects.create(
            teacher_id=subject_teacher.teacher_id,
            section_id=section.section_id,
            mapping_id=mapping.mapping_id,
        )
        messages.success(
            request,
            f'✅ {subject_teacher.full_name} assigned to {mapping.subject_name} in {section.section_name}.'
        )
        return redirect('adviser_subject_assignments')

    # Includes the adviser themself, so they can assign themselves to teach
    # a subject in their own section.
    subject_teachers = Teacher.objects.filter(
        school_profile_id=teacher.school_profile_id,
        role__in=(Teacher.ROLE_ADVISER, Teacher.ROLE_SUBJECT_TEACHER),
    ).order_by('full_name')

    # A blank mapping.strand applies to every strand at that grade level, and
    # a blank section.strand (no strand distinction for this section at all)
    # matches every mapping regardless of its strand.
    mappings = SubjectMapping.objects.filter(
        school_profile_id=teacher.school_profile_id,
        grade_level=section.grade_level,
        is_active=True,
    )
    if section.strand:
        mappings = mappings.filter(Q(strand='') | Q(strand=section.strand))
    mappings = mappings.order_by('subject_name')

    # Narrow the teacher dropdown to whoever's actually relevant to this
    # section's grade level - an adviser's own section grade level, or a
    # subject teacher's existing assignments' grade level(s). A teacher
    # with neither yet (a brand new account) has no established grade
    # level to filter on, so they stay offered everywhere - otherwise
    # they could never receive a first assignment at all.
    candidate_ids = [t.teacher_id for t in subject_teachers]
    adviser_grade_level_by_teacher = {
        s.adviser_id: s.grade_level
        for s in Section.objects.filter(school_profile_id=teacher.school_profile_id, adviser_id__in=candidate_ids)
    }
    candidate_assignments = TeacherSubjectAssignment.objects.filter(teacher_id__in=candidate_ids)
    assignment_mapping_ids = {a.mapping_id for a in candidate_assignments}
    grade_level_by_mapping = dict(
        SubjectMapping.objects.filter(mapping_id__in=assignment_mapping_ids).values_list('mapping_id', 'grade_level')
    )
    assignment_grade_levels_by_teacher = {}
    for a in candidate_assignments:
        grade_level = grade_level_by_mapping.get(a.mapping_id)
        if grade_level:
            assignment_grade_levels_by_teacher.setdefault(a.teacher_id, set()).add(grade_level)

    def _relevant_to_section(t):
        known_levels = assignment_grade_levels_by_teacher.get(t.teacher_id, set())
        if t.teacher_id in adviser_grade_level_by_teacher:
            known_levels = known_levels | {adviser_grade_level_by_teacher[t.teacher_id]}
        return not known_levels or section.grade_level in known_levels

    subject_teachers = [t for t in subject_teachers if _relevant_to_section(t)]

    teachers_by_id = {t.teacher_id: t for t in subject_teachers}
    mappings_by_id = {m.mapping_id: m for m in mappings}

    assignments = TeacherSubjectAssignment.objects.filter(
        section_id=section.section_id
    ).order_by('assignment_id')

    assignment_rows = [
        {
            'assignment': a,
            'teacher': teachers_by_id.get(a.teacher_id)
                or Teacher.objects.filter(teacher_id=a.teacher_id).first(),
            'mapping': mappings_by_id.get(a.mapping_id)
                or SubjectMapping.objects.filter(mapping_id=a.mapping_id).first(),
        }
        for a in assignments
    ]

    return render(request, 'school/adviser_assignments.html', {
        'section': section,
        'subject_teachers': subject_teachers,
        'mappings': mappings,
        'assignment_rows': assignment_rows,
    })


@login_required
def remove_adviser_subject_assignment(request, assignment_id):
    if request.method != 'POST':
        return redirect('adviser_subject_assignments')

    teacher, section, error = _get_adviser_with_section(request)
    if error:
        return error

    # Scoped to the adviser's own section - even a guessed assignment_id from
    # another section won't match this filter.
    assignment = TeacherSubjectAssignment.objects.filter(
        assignment_id=assignment_id, section_id=section.section_id
    ).first()
    if not assignment:
        messages.error(request, 'Assignment not found in your section.')
        return redirect('adviser_subject_assignments')

    assignment_teacher = Teacher.objects.filter(teacher_id=assignment.teacher_id).first()
    assignment.delete()
    name = assignment_teacher.full_name if assignment_teacher else 'Teacher'
    messages.success(request, f'Assignment removed from {name}.')
    return redirect('adviser_subject_assignments')


# ---------------------------------------------------------------------------
# Class Scheduling
# ---------------------------------------------------------------------------

def _require_admin_teacher_json(request):
    """Same gate as _get_school_admin_teacher, but returns None instead of
    an HTML redirect - the caller is a JSON endpoint, so a redirect
    response would just look like a failed fetch() to the browser."""
    try:
        teacher = Teacher.objects.get(username=request.user.username)
    except Teacher.DoesNotExist:
        return None
    is_officer = teacher.role == Teacher.ROLE_NON_TEACHING and teacher.is_officer
    if teacher.role not in (Teacher.ROLE_REGISTRAR, Teacher.ROLE_PRINCIPAL) and not is_officer:
        return None
    return teacher


@login_required
def school_scheduling(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    if not teacher.school_profile_id:
        messages.warning(request, 'Your account has no school assigned yet.')
        return render(request, 'school/scheduling.html', {'sections': [], 'teachers': []})

    school_profile_id = teacher.school_profile_id
    scheduler.ensure_time_slots(school_profile_id)

    sections = Section.objects.filter(school_profile_id=school_profile_id).order_by(
        'grade_level', 'strand', 'section_name'
    )
    subject_teachers = Teacher.objects.filter(
        school_profile_id=school_profile_id,
        role__in=(Teacher.ROLE_SUBJECT_TEACHER, Teacher.ROLE_ADVISER),
    ).order_by('full_name')

    if not sections:
        messages.info(request, 'Add a section first before building a schedule.')
        return render(request, 'school/scheduling.html', {
            'sections': [], 'teachers': subject_teachers,
        })

    view_mode = request.GET.get('view') if request.GET.get('view') in ('section', 'teacher') else 'section'
    slots = list(TimeSlot.objects.filter(school_profile_id=school_profile_id).order_by('slot_order'))
    days = [d for d, _ in ScheduleEntry.DAY_CHOICES]

    context = {
        'sections': sections,
        'teachers': subject_teachers,
        'days': days,
        'view_mode': view_mode,
    }

    if view_mode == 'teacher':
        teacher_id = request.GET.get('teacher_id')
        selected_teacher = subject_teachers.filter(teacher_id=teacher_id).first() or subject_teachers.first()
        context['selected_teacher'] = selected_teacher

        grid = {}
        if selected_teacher:
            sections_by_id = {s.section_id: s for s in sections}
            mappings_by_id = {m.mapping_id: m for m in SubjectMapping.objects.filter(school_profile_id=school_profile_id)}
            entries = ScheduleEntry.objects.filter(
                school_profile_id=school_profile_id, teacher_id=selected_teacher.teacher_id
            )
            for e in entries:
                mapping = mappings_by_id.get(e.mapping_id)
                section = sections_by_id.get(e.section_id)
                grid[(e.timeslot_id, e.day_of_week)] = {
                    'subject': mapping.subject_name if mapping else '',
                    'section': section.section_name if section else '',
                }

        slot_rows = []
        for slot in slots:
            if slot.is_break:
                slot_rows.append({'slot': slot, 'is_break': True})
                continue
            cells = [{'day': day, 'info': grid.get((slot.timeslot_id, day))} for day in days]
            slot_rows.append({'slot': slot, 'is_break': False, 'cells': cells})
        context['slot_rows'] = slot_rows

    else:
        section_id = request.GET.get('section_id')
        selected_section = sections.filter(section_id=section_id).first() or sections.first()
        context['selected_section'] = selected_section

        # Subjects offered for "+ Add to Roster" are narrowed to this
        # section's grade level (and strand, when the subject specifies
        # one) - the same match school_assignments enforces on submit, so
        # picking from this list can't fail that check.
        context['mappings'] = SubjectMapping.objects.filter(
            school_profile_id=school_profile_id, is_active=True, grade_level=selected_section.grade_level
        ).filter(Q(strand='') | Q(strand=selected_section.strand)).order_by('subject_name')

        assignments = list(TeacherSubjectAssignment.objects.filter(section_id=selected_section.section_id))
        teachers_by_id = {t.teacher_id: t for t in subject_teachers}
        mappings_by_id = {m.mapping_id: m for m in SubjectMapping.objects.filter(school_profile_id=school_profile_id)}
        requirements = {
            r.assignment_id: r.sessions_per_week
            for r in ScheduleRequirement.objects.filter(
                assignment_id__in=[a.assignment_id for a in assignments]
            )
        }

        entries = list(ScheduleEntry.objects.filter(section_id=selected_section.section_id))
        placed_count = defaultdict(int)
        grid = {}
        for e in entries:
            grid[(e.timeslot_id, e.day_of_week)] = e.assignment_id
            placed_count[e.assignment_id] += 1

        roster = []
        for a in assignments:
            t = teachers_by_id.get(a.teacher_id)
            m = mappings_by_id.get(a.mapping_id)
            needed = requirements.get(a.assignment_id, scheduler.DEFAULT_SESSIONS_PER_WEEK)
            roster.append({
                'assignment': a,
                'teacher': t,
                'mapping': m,
                'needed': needed,
                'placed': placed_count.get(a.assignment_id, 0),
                'label': f"{m.subject_name if m else '(subject removed)'} — {t.full_name if t else '(teacher removed)'}",
            })
        context['roster'] = roster

        slot_rows = []
        for slot in slots:
            if slot.is_break:
                slot_rows.append({'slot': slot, 'is_break': True})
                continue
            cells = [{'day': day, 'selected_id': grid.get((slot.timeslot_id, day))} for day in days]
            slot_rows.append({'slot': slot, 'is_break': False, 'cells': cells})
        context['slot_rows'] = slot_rows

    return render(request, 'school/scheduling.html', context)


@require_POST
def save_schedule_cell(request):
    teacher = _require_admin_teacher_json(request)
    if not teacher:
        return JsonResponse({'ok': False, 'error': 'Not authorized.'}, status=403)

    try:
        payload = json.loads(request.body)
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'Bad request.'}, status=400)

    day = payload.get('day')
    section_id = payload.get('section_id')
    timeslot_id = payload.get('timeslot_id')
    assignment_id = payload.get('assignment_id') or None

    if day not in dict(ScheduleEntry.DAY_CHOICES):
        return JsonResponse({'ok': False, 'error': 'Invalid day.'}, status=400)

    section = Section.objects.filter(section_id=section_id, school_profile_id=teacher.school_profile_id).first()
    timeslot = TimeSlot.objects.filter(
        timeslot_id=timeslot_id, school_profile_id=teacher.school_profile_id, is_break=False
    ).first()
    if not section or not timeslot:
        return JsonResponse({'ok': False, 'error': 'Section or time slot not found.'}, status=404)

    with transaction.atomic():
        ScheduleEntry.objects.filter(
            section_id=section.section_id, day_of_week=day, timeslot_id=timeslot.timeslot_id
        ).delete()

        if not assignment_id:
            return JsonResponse({'ok': True, 'cleared': True})

        assignment = TeacherSubjectAssignment.objects.filter(
            assignment_id=assignment_id, section_id=section.section_id
        ).first()
        if not assignment:
            return JsonResponse({'ok': False, 'error': 'That assignment is not part of this section.'}, status=404)

        conflict = ScheduleEntry.objects.filter(
            teacher_id=assignment.teacher_id, day_of_week=day, timeslot_id=timeslot.timeslot_id
        ).exclude(section_id=section.section_id).first()
        if conflict:
            conflict_teacher = Teacher.objects.filter(teacher_id=assignment.teacher_id).first()
            conflict_section = Section.objects.filter(section_id=conflict.section_id).first()
            return JsonResponse({
                'ok': False,
                'error': (
                    f"{conflict_teacher.full_name if conflict_teacher else 'This teacher'} is already "
                    f"teaching {conflict_section.section_name if conflict_section else 'another section'} "
                    f"at this time on {day}."
                ),
            }, status=409)

        try:
            ScheduleEntry.objects.create(
                school_profile_id=teacher.school_profile_id,
                assignment_id=assignment.assignment_id,
                teacher_id=assignment.teacher_id,
                section_id=section.section_id,
                mapping_id=assignment.mapping_id,
                day_of_week=day,
                timeslot_id=timeslot.timeslot_id,
            )
        except IntegrityError:
            return JsonResponse({'ok': False, 'error': 'That slot was just taken. Try again.'}, status=409)

    subject = SubjectMapping.objects.filter(mapping_id=assignment.mapping_id).first()
    placed_teacher = Teacher.objects.filter(teacher_id=assignment.teacher_id).first()
    return JsonResponse({
        'ok': True,
        'subject_name': subject.subject_name if subject else '',
        'teacher_name': placed_teacher.full_name if placed_teacher else '',
    })


@login_required
@require_POST
def add_schedule_roster(request):
    """The 'holistic' entry point: pick a teacher, a subject, and how many
    times/week - nothing about day or time. Creates (or reuses) the
    TeacherSubjectAssignment + ScheduleRequirement, then immediately runs
    the auto-scheduler so the new sessions land on the grid without a
    separate Auto-Generate click."""
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    section_id = request.POST.get('section_id', '')
    redirect_url = reverse('school_scheduling') + f'?section_id={section_id}'

    subject_teacher_id = request.POST.get('teacher_id', '').strip()
    mapping_id = request.POST.get('mapping_id', '').strip()
    sessions_per_week = request.POST.get('sessions_per_week', '').strip()

    if not (subject_teacher_id and mapping_id):
        messages.error(request, 'Pick a teacher and a subject.')
        return redirect(redirect_url)

    section = Section.objects.filter(section_id=section_id, school_profile_id=teacher.school_profile_id).first()
    subject_teacher = Teacher.objects.filter(
        teacher_id=subject_teacher_id,
        school_profile_id=teacher.school_profile_id,
        role__in=(Teacher.ROLE_SUBJECT_TEACHER, Teacher.ROLE_ADVISER),
    ).first()
    mapping = SubjectMapping.objects.filter(
        mapping_id=mapping_id, school_profile_id=teacher.school_profile_id
    ).first()

    if not (section and subject_teacher and mapping):
        messages.error(request, 'Section, teacher, or subject not found in this school.')
        return redirect(redirect_url)

    if not _mapping_matches_section(mapping, section):
        messages.error(
            request,
            f"{mapping.subject_name} (Grade {mapping.grade_level} {mapping.strand}) doesn't match "
            f"{section.grade_level}-{section.strand}-{section.section_name}."
        )
        return redirect(redirect_url)

    if not sessions_per_week.isdigit() or int(sessions_per_week) < 1:
        sessions_per_week = str(scheduler.DEFAULT_SESSIONS_PER_WEEK)

    assignment = TeacherSubjectAssignment.objects.filter(
        teacher_id=subject_teacher.teacher_id, section_id=section.section_id, mapping_id=mapping.mapping_id,
    ).first()
    if not assignment:
        assignment = TeacherSubjectAssignment.objects.create(
            teacher_id=subject_teacher.teacher_id, section_id=section.section_id, mapping_id=mapping.mapping_id,
        )

    ScheduleRequirement.objects.update_or_create(
        assignment_id=assignment.assignment_id,
        defaults={'sessions_per_week': int(sessions_per_week)},
    )

    scheduler.ensure_time_slots(teacher.school_profile_id)
    result = scheduler.generate_schedule(teacher.school_profile_id)

    messages.success(
        request,
        f'✅ {subject_teacher.full_name} — {mapping.subject_name} ({sessions_per_week}x/week) added. '
        f"Auto-placed {result['created']} session(s) across the school."
    )
    for item in result['unresolved']:
        messages.warning(
            request,
            f"Couldn't fit {item['still_needed']} more session(s) of {item['subject']} "
            f"({item['teacher']} · {item['section']}) anywhere conflict-free."
        )

    return redirect(redirect_url)


@login_required
@require_POST
def auto_generate_schedule(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    if not teacher.school_profile_id:
        messages.warning(request, 'Your account has no school assigned yet.')
        return redirect('school_scheduling')

    scheduler.ensure_time_slots(teacher.school_profile_id)
    result = scheduler.generate_schedule(teacher.school_profile_id)

    if result['created']:
        messages.success(request, f"✅ Auto-generated {result['created']} class session(s).")
    else:
        messages.info(
            request,
            'Nothing to auto-generate — every assignment is already fully scheduled '
            '(or there is no teacher/subject/section roster yet — set that up under Manage Assignments).'
        )

    for item in result['unresolved']:
        messages.warning(
            request,
            f"Couldn't fit {item['still_needed']} more session(s) of {item['subject']} "
            f"({item['teacher']} · {item['section']}) anywhere conflict-free — "
            f"free up a slot manually and re-run, or clear a conflicting entry."
        )

    query = request.POST.get('return_qs', '')
    url = reverse('school_scheduling')
    return redirect(url + ('?' + query if query else ''))


@login_required
@require_POST
def clear_section_schedule(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    section_id = request.POST.get('section_id', '')
    section = Section.objects.filter(section_id=section_id, school_profile_id=teacher.school_profile_id).first()
    if not section:
        messages.error(request, 'Section not found.')
        return redirect('school_scheduling')

    deleted, _detail = ScheduleEntry.objects.filter(section_id=section.section_id).delete()
    messages.success(request, f'Cleared {deleted} scheduled session(s) for {section.section_name}.')
    return redirect(reverse('school_scheduling') + f'?section_id={section.section_id}')


@login_required
@require_POST
def save_schedule_requirement(request):
    teacher, error = _get_school_admin_teacher(request)
    if error:
        return error

    assignment_id = request.POST.get('assignment_id', '')
    section_id = request.POST.get('section_id', '')
    sessions_per_week = request.POST.get('sessions_per_week', '').strip()

    assignment = TeacherSubjectAssignment.objects.filter(assignment_id=assignment_id).first()
    valid_section = None
    if assignment:
        valid_section = Section.objects.filter(
            section_id=assignment.section_id, school_profile_id=teacher.school_profile_id
        ).first()

    if not assignment or not valid_section or not sessions_per_week.isdigit() or int(sessions_per_week) < 1:
        messages.error(request, 'Could not update that requirement.')
    else:
        ScheduleRequirement.objects.update_or_create(
            assignment_id=assignment.assignment_id,
            defaults={'sessions_per_week': int(sessions_per_week)},
        )
        messages.success(request, 'Updated sessions/week.')

    return redirect(reverse('school_scheduling') + f'?section_id={section_id}')
