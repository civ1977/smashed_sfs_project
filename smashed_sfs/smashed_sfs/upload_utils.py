"""Shared CSV/Excel row reader for the student and grades upload forms.

Both upload flows accept a file, build an editable preview from it, and
only actually save once the teacher submits that preview form - so this
only needs to turn whatever was uploaded into plain string rows; the
existing preview/save code on each side is unchanged.
"""
import csv
import io
from datetime import date, datetime

import openpyxl


class UploadFileError(Exception):
    """Raised when an uploaded file can't be read as CSV or Excel, with a
    message safe to show directly to the teacher."""


# Class lists/gradesheets are at most a few hundred rows - genuinely that
# size in .xlsx/.csv is well under 1 MB, so this is a generous cap against
# an oversized file tying up openpyxl.load_workbook, not a realistic limit
# for the actual use case.
MAX_UPLOAD_BYTES = 10 * 1024 * 1024


def _normalize_cell(value):
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        # Matches convert_date()'s own '%Y-%m-%d' format, so a real Excel
        # date cell round-trips through the editable preview form without
        # the teacher needing to retype it.
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_upload_rows(uploaded_file):
    """Reads an uploaded .csv or .xlsx file into a list of rows, each a
    list of plain strings - callers treat the result identically
    regardless of which format was actually uploaded. Blank rows are
    dropped from both formats the same way a blank CSV line already was.
    """
    if uploaded_file.size > MAX_UPLOAD_BYTES:
        raise UploadFileError(
            f'That file is too large ({uploaded_file.size // (1024 * 1024)} MB) - '
            f'the limit is {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.'
        )

    name = (uploaded_file.name or '').lower()

    if name.endswith('.xlsx') or name.endswith('.xlsm'):
        try:
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
        except Exception:
            raise UploadFileError(
                'This file could not be read as an Excel workbook. Make sure it is a real '
                '.xlsx file (not a renamed .csv, and not a legacy .xls file).'
            )
        sheet = workbook.active
        rows = []
        for raw_row in sheet.iter_rows(values_only=True):
            row = [_normalize_cell(v) for v in raw_row]
            if any(cell != '' for cell in row):
                rows.append(row)
        return rows

    if name.endswith('.xls'):
        raise UploadFileError(
            'Legacy .xls files are not supported - please save as .xlsx or .csv and try again.'
        )

    decoded_file = uploaded_file.read().decode('utf-8-sig')
    reader = csv.reader(io.StringIO(decoded_file), delimiter=',', quotechar='"')
    return [row for row in reader if row]
